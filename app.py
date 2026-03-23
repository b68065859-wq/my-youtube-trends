"""
╔══════════════════════════════════════════════════════╗
║         ABS VIRAL 777 — Professional Edition         ║
║         ViewStats-style YouTube Analytics            ║
╚══════════════════════════════════════════════════════╝

pip install streamlit google-api-python-client pandas pyTelegramBotAPI plotly

Streamlit Cloud → Settings → Secrets:
[general]
BOT_TOKEN         = "123456:ABC..."
ADMIN_CHAT_ID     = "123456789"
PAYME_MERCHANT    = "your_id"
CLICK_SERVICE     = "your_id"
CLICK_MERCHANT    = "your_id"
TELEGRAM_BOT_LINK = "https://t.me/your_bot"
APP_URL           = "https://my-youtube-trends-nk9rk2csy8hjj6dvhfbgku.streamlit.app"
"""

import streamlit as st
import googleapiclient.discovery
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, timedelta
import re, json, os, base64, uuid, threading, random, string, time, html as html_mod
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import telebot
    BOT_AVAILABLE = True
except:
    BOT_AVAILABLE = False

# ══════════════════════════════════════════
# КОНФИГ
# ══════════════════════════════════════════
FREE_TRIAL_LIMIT   = 3
SUBSCRIPTION_PRICE = 50000
SUBSCRIPTION_DAYS  = 30
ADMIN_DB           = {"baho123": "qWe83664323546"}
DB_FILE            = "/tmp/viral777_db.json"

def cfg(key, default=""):
    try:    return st.secrets["general"][key]
    except: return os.environ.get(key, default)

BOT_TOKEN         = cfg("BOT_TOKEN")
ADMIN_CHAT_ID     = cfg("ADMIN_CHAT_ID")
PAYME_MERCHANT    = cfg("PAYME_MERCHANT")
CLICK_SERVICE     = cfg("CLICK_SERVICE")
CLICK_MERCHANT    = cfg("CLICK_MERCHANT")
TELEGRAM_BOT_LINK = cfg("TELEGRAM_BOT_LINK", "https://t.me/your_bot")

# ══════════════════════════════════════════
# PAGE CONFIG
# ══════════════════════════════════════════
st.set_page_config(
    page_title="Viral 777 — YouTube Analytics",
    page_icon="🔥",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={"Get Help":None,"Report a bug":None,"About":None}
)

# Dark/Light mode
def apply_theme():
    if not st.session_state.get("dark_mode", True):
        st.markdown("""<style>
        html, body, .stApp, .main, .block-container,
        [data-testid="stAppViewContainer"],
        [data-testid="stAppViewBlockContainer"] {
            background-color: #f0f2f8 !important; color: #1a1a2e !important;
        }
        header[data-testid="stHeader"] { background: #f0f2f8 !important; }
        [data-testid="stSidebar"], [data-testid="stSidebar"] > div {
            background: linear-gradient(180deg, #e8eaf6 0%, #dde1f5 100%) !important;
            border-right: 1px solid #c5cae9 !important;
        }
        [data-testid="stSidebar"] * { color: #1a1a2e !important; }
        [data-testid="stSidebar"] input { background: #ffffff !important; color: #1a1a2e !important; border-color: #c5cae9 !important; }
        [data-testid="stSidebar"] [data-baseweb="select"] > div { background: #ffffff !important; color: #1a1a2e !important; border-color: #c5cae9 !important; }
        .stat-card { background: #ffffff !important; border-color: #c5cae9 !important; box-shadow: 0 2px 16px rgba(63,81,181,0.1) !important; }
        .stat-card.purple { border-top: 3px solid #3f51b5 !important; }
        .stat-card.red    { border-top: 3px solid #e53935 !important; }
        .stat-card.gold   { border-top: 3px solid #ffa000 !important; }
        .stat-card.green  { border-top: 3px solid #43a047 !important; }
        .stat-label { color: #7986cb !important; }
        .stat-value { color: #1a1a2e !important; }
        .stat-sub   { color: #9fa8da !important; }
        .video-card { background: #ffffff !important; border-color: #c5cae9 !important; }
        .vc-title { color: #1a1a2e !important; }
        .vc-channel { color: #7986cb !important; }
        .vc-stat-val { color: #1a1a2e !important; }
        .vc-stat-lbl { color: #9fa8da !important; }
        .vs-logo { background: linear-gradient(135deg,#3f51b5,#e91e63) !important; -webkit-background-clip:text !important; -webkit-text-fill-color:transparent !important; }
        .section-title { color: #1a1a2e !important; }
        .viral-score-box { background: #ffffff !important; border-color: #c5cae9 !important; }
        [data-baseweb="tab-list"] { background: #e8eaf6 !important; border-bottom: 1px solid #c5cae9 !important; }
        [data-baseweb="tab"] { color: #7986cb !important; background: transparent !important; }
        [aria-selected="true"][data-baseweb="tab"] { color: #3f51b5 !important; border-bottom: 2px solid #3f51b5 !important; background: #dde1f5 !important; }
        input, textarea { background: #ffffff !important; color: #1a1a2e !important; border-color: #c5cae9 !important; }
        [data-baseweb="select"] > div { background: #ffffff !important; color: #1a1a2e !important; border-color: #c5cae9 !important; }
        [data-baseweb="select"] div, [data-baseweb="select"] span, [data-baseweb="select"] p { color: #1a1a2e !important; }
        [data-baseweb="select"] svg { fill: #3f51b5 !important; }
        [data-baseweb="option"] { background: #ffffff !important; color: #1a1a2e !important; }
        [data-baseweb="option"]:hover { background: #e8eaf6 !important; }
        [data-baseweb="popover"], [data-baseweb="popover"] * { background: #ffffff !important; color: #1a1a2e !important; }
        [data-baseweb="menu"], [data-baseweb="menu"] * { background: #ffffff !important; color: #1a1a2e !important; }
        li[role="option"], li[role="option"] * { background: #ffffff !important; color: #1a1a2e !important; }
        .stButton > button { background: linear-gradient(135deg,#3f51b5,#5c6bc0) !important; color: #fff !important; }
        [data-testid="stAlert"] { background: #e8eaf6 !important; color: #1a1a2e !important; border-color: #c5cae9 !important; }
        [data-testid="stDataFrame"] * { background-color: #ffffff !important; color: #1a1a2e !important; }
        [data-testid="stDataFrame"] th, [data-testid="stDataFrame"] [role="columnheader"] { background-color: #e8eaf6 !important; color: #3f51b5 !important; }
        [data-testid="stDataFrame"] td, [data-testid="stDataFrame"] [role="gridcell"] { background-color: #ffffff !important; color: #1a1a2e !important; }
        [data-testid="stMarkdownContainer"] p { color: #3a3a5e !important; }
        hr { border-color: #c5cae9 !important; }
        [data-testid="stExpander"] { background: #ffffff !important; border-color: #c5cae9 !important; }
        [data-testid="stExpander"] summary { color: #1a1a2e !important; }
        </style>""", unsafe_allow_html=True)

# ══════════════════════════════════════════
# CSS
# ══════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap');
*, *::before, *::after { box-sizing: border-box; margin: 0; }

/* MANAGE APP YASHIRISH */
#MainMenu { visibility: hidden !important; display: none !important; }
header[data-testid="stHeader"] {
    background: #0a0a0f !important;
    height: 0 !important; min-height: 0 !important;
    padding: 0 !important; overflow: hidden !important;
}
[data-testid="stToolbar"]         { display: none !important; }
[data-testid="stStatusWidget"]    { display: none !important; }
[data-testid="stAppDeployButton"] { display: none !important; }
[data-testid="manage-app-button"] { display: none !important; }
[data-testid="stBottom"]          { display: none !important; }
[class*="viewerBadge"]    { display: none !important; }
[class*="StatusWidget"]   { display: none !important; }
[class*="toolbarActions"] { display: none !important; }
[title="Manage app"]      { display: none !important; }
#stDecoration             { display: none !important; }
footer, footer *          { display: none !important; }
.stApp > header           { display: none !important; }
section[data-testid="stSidebarNav"] { display: none !important; }
/* Pastki o'ng badge — hammasi yashirish */
div[class*="badge"], div[class*="Badge"] { display: none !important; }
img[alt="Streamlit"] { display: none !important; }
/* Streamlit community cloud belgilari */
[class*="ActionButton"]   { display: none !important; }
[class*="action-button"]  { display: none !important; }
[data-testid="stActionButton"] { display: none !important; }
a[href*="streamlit.io"]   { display: none !important; }
a[href*="share.streamlit"] { display: none !important; }
/* O'ng pastki fixed elementlar */
.stApp > div > div:last-child > div:last-child > div:last-child { display: none !important; }
iframe[title*="streamlit"] { display: none !important; }
/* Har qanday fixed pastki element */
div[style*="position: fixed"][style*="bottom"] { display: none !important; }
div[style*="position:fixed"][style*="bottom"]  { display: none !important; }

/* GLOBAL */
html, body, .stApp, .main, .block-container {
    background-color: #0a0a0f !important;
    color: #e8e8f0 !important;
    font-family: 'Inter', sans-serif !important;
}
.block-container {
    padding: 1.5rem 2rem !important;
    max-width: 1400px !important;
    padding-top: 1rem !important;
}

/* SIDEBAR */
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #0f0f1a 0%, #0a0a12 100%) !important;
    border-right: 1px solid #1e1e2e !important;
}
[data-testid="stSidebar"] * { color: #e8e8f0 !important; }
[data-testid="stSidebar"] input {
    background: #1a1a2e !important;
    border: 1px solid #2a2a4a !important;
    color: #fff !important;
    border-radius: 8px !important;
}
[data-testid="stSidebar"] [data-baseweb="select"] > div {
    background: #1a1a2e !important;
    border: 1px solid #2a2a4a !important;
    border-radius: 8px !important;
}

/* INPUTS */
input, textarea, [data-baseweb="input"] input {
    background: #1a1a2e !important;
    color: #ffffff !important;
    border: 1px solid #2a2a4a !important;
    border-radius: 8px !important;
}

/* SELECTBOX */
[data-baseweb="select"] { color: #ffffff !important; }
[data-baseweb="select"] > div {
    background: #1a1a2e !important; color: #ffffff !important;
    border: 1px solid #2a2a4a !important; border-radius: 8px !important;
}
[data-baseweb="select"] div, [data-baseweb="select"] span,
[data-baseweb="select"] p, [data-baseweb="select"] input,
[data-baseweb="select"] [class*="placeholder"],
[data-baseweb="select"] [class*="singleValue"],
[data-baseweb="select"] [class*="ValueContainer"] { color: #ffffff !important; }
[data-baseweb="select"] svg { fill: #ffffff !important; }
[data-baseweb="option"] { background: #1a1a2e !important; color: #ffffff !important; }
[data-baseweb="option"]:hover,
[data-baseweb="option"][aria-selected="true"] { background: #2a2a4a !important; color: #ffffff !important; }
[data-baseweb="popover"] { background: #1a1a2e !important; border: 1px solid #2a2a4a !important; }
[data-baseweb="popover"] * { background: #1a1a2e !important; color: #ffffff !important; }
[data-baseweb="menu"] { background: #1a1a2e !important; }
[data-baseweb="menu"] * { color: #ffffff !important; }
li[role="option"], li[role="option"] * { color: #ffffff !important; background: #1a1a2e !important; }
li[role="option"]:hover { background: #2a2a4a !important; }
[data-testid="stSelectbox"] label { color: #e8e8f0 !important; }

/* RADIO */
[role="radiogroup"] label, [role="radiogroup"] p,
div[data-testid="stRadio"] label, div[data-testid="stRadio"] p { color: #e8e8f0 !important; }

/* SLIDER */
[data-testid="stSlider"] p, [data-testid="stSlider"] span,
[data-testid="stSlider"] label { color: #e8e8f0 !important; }

/* TABS */
[data-baseweb="tab-list"] {
    background: transparent !important;
    border-bottom: 1px solid #1e1e2e !important;
    gap: 4px !important;
}
[data-baseweb="tab"] {
    color: #666688 !important; background: transparent !important;
    border-radius: 8px 8px 0 0 !important;
    padding: 10px 20px !important;
    font-weight: 600 !important; font-size: 14px !important;
    transition: all 0.2s !important;
}
[data-baseweb="tab"]:hover { color: #aaaacc !important; background: #1a1a2e !important; }
[aria-selected="true"][data-baseweb="tab"] {
    color: #ffffff !important; background: #1a1a2e !important;
    border-bottom: 2px solid #6c63ff !important;
}

/* BUTTONS */
.stButton > button {
    background: linear-gradient(135deg, #6c63ff 0%, #5a54e8 100%) !important;
    color: #ffffff !important; border-radius: 10px !important;
    font-weight: 700 !important; border: none !important;
    height: 2.8em !important; width: 100% !important;
    transition: all 0.25s !important; font-size: 14px !important;
}
.stButton > button:hover {
    transform: translateY(-1px) !important;
    box-shadow: 0 6px 20px rgba(108,99,255,0.45) !important;
}
.stButton > button:disabled { background: #1e1e2e !important; color: #444466 !important; }

/* DATAFRAME */
[data-testid="stDataFrame"] { border-radius: 12px !important; overflow: hidden !important; }
[data-testid="stDataFrame"] * {
    background-color: #0f0f1a !important; color: #e8e8f0 !important; border-color: #1e1e2e !important;
}
[data-testid="stDataFrame"] th,
[data-testid="stDataFrame"] [role="columnheader"] {
    background-color: #1a1a2e !important; color: #9c93ff !important; font-weight: 700 !important;
}
[data-testid="stDataFrame"] td,
[data-testid="stDataFrame"] [role="gridcell"] {
    background-color: #0f0f1a !important; color: #e8e8f0 !important;
}
[data-testid="stDataFrame"] tr:hover td { background-color: #1a1a2e !important; }
[data-testid="stDataFrame"] iframe { background: #0f0f1a !important; }
.dvn-scroller { background: #0f0f1a !important; }

/* ALERTS */
[data-testid="stAlert"] { background: #1a1a2e !important; border-radius: 10px !important; color: #e8e8f0 !important; border: 1px solid #2a2a4a !important; }

/* EXPANDER */
[data-testid="stExpander"] { background: #0f0f1a !important; border: 1px solid #1e1e2e !important; border-radius: 12px !important; }
[data-testid="stExpander"] summary { color: #e8e8f0 !important; }
[data-testid="stExpander"] p { color: #aaaacc !important; }

/* MARKDOWN */
[data-testid="stMarkdownContainer"] p { color: #ccccdd !important; }
.stCaption { color: #666688 !important; }
hr { border-color: #1e1e2e !important; }

/* CUSTOM COMPONENTS */
.vs-header { display: flex; align-items: center; gap: 12px; padding: 4px 0 16px; border-bottom: 1px solid #1e1e2e; margin-bottom: 16px; }
.vs-logo { font-size: 22px; font-weight: 900; background: linear-gradient(135deg, #6c63ff, #ff6b6b); -webkit-background-clip: text; -webkit-text-fill-color: transparent; }
.vs-tagline { font-size: 11px; color: #555577; font-weight: 500; }

.stat-card { background: linear-gradient(135deg, #0f0f1a 0%, #12121f 100%); border: 1px solid #1e1e2e; border-radius: 16px; padding: 20px 24px; transition: all 0.3s; position: relative; overflow: hidden; }
.stat-card:hover { border-color: #3a3a5e; transform: translateY(-2px); box-shadow: 0 8px 32px rgba(0,0,0,0.4); }
.stat-card::before { content: ''; position: absolute; top: 0; left: 0; right: 0; height: 2px; }
.stat-card.purple::before { background: linear-gradient(90deg, #6c63ff, #9c93ff); }
.stat-card.red::before    { background: linear-gradient(90deg, #ff4757, #ff6b81); }
.stat-card.green::before  { background: linear-gradient(90deg, #2ed573, #7bed9f); }
.stat-card.gold::before   { background: linear-gradient(90deg, #ffa502, #ffcc02); }
.stat-label { font-size: 11px; font-weight: 600; color: #555577; text-transform: uppercase; letter-spacing: 1px; margin-bottom: 8px; }
.stat-value { font-size: 28px; font-weight: 900; color: #ffffff; line-height: 1; }
.stat-sub   { font-size: 12px; color: #555577; margin-top: 6px; }
.stat-icon  { position: absolute; right: 20px; top: 20px; font-size: 28px; opacity: 0.15; }

.video-card { background: #0f0f1a; border: 1px solid #1e1e2e; border-radius: 16px; padding: 0; overflow: hidden; transition: all 0.3s; }
.video-card:hover { border-color: #3a3a5e; transform: translateY(-3px); box-shadow: 0 12px 40px rgba(0,0,0,0.5); }
.vc-thumbnail { width: 100%; aspect-ratio: 16/9; object-fit: cover; }
.vc-body { padding: 14px 16px; }
.vc-title { font-size: 13px; font-weight: 600; color: #e8e8f0; line-height: 1.4; margin-bottom: 8px; display: -webkit-box; -webkit-line-clamp: 2; -webkit-box-orient: vertical; overflow: hidden; }
.vc-channel { font-size: 11px; color: #555577; margin-bottom: 10px; }
.vc-stats { display: flex; gap: 12px; }
.vc-stat { display: flex; flex-direction: column; }
.vc-stat-val { font-size: 15px; font-weight: 800; color: #ffffff; }
.vc-stat-lbl { font-size: 10px; color: #444466; text-transform: uppercase; }
.vc-badge { display: inline-block; padding: 3px 8px; border-radius: 6px; font-size: 11px; font-weight: 700; margin-bottom: 8px; }
.vc-badge.fire { background: rgba(255,71,87,0.15); color: #ff4757; border: 1px solid rgba(255,71,87,0.3); }
.vc-badge.hot  { background: rgba(255,165,2,0.15);  color: #ffa502; border: 1px solid rgba(255,165,2,0.3); }
.vc-badge.ok   { background: rgba(46,213,115,0.15); color: #2ed573; border: 1px solid rgba(46,213,115,0.3); }

.sub-box { background: linear-gradient(135deg, #0f0f1a, #12121f); border: 1px solid #2a2a4a; border-radius: 20px; padding: 48px 40px; text-align: center; margin: 20px 0; position: relative; overflow: hidden; }
.sub-box::before { content: ''; position: absolute; top: 0; left: 0; right: 0; height: 3px; background: linear-gradient(90deg, #6c63ff, #ff6b6b, #ffa502); }
.act-box { background: #0f0f1a; border: 1px solid #2a2a4a; border-radius: 16px; padding: 28px; text-align: center; margin: 16px 0; }
.tg-pay-btn { display: inline-flex; align-items: center; gap: 10px; background: linear-gradient(135deg, #229ED9, #1a8fc2); color: #fff !important; padding: 16px 40px; border-radius: 12px; font-weight: 800; font-size: 16px; text-decoration: none !important; transition: 0.3s; margin: 12px 0; }

.badge-trial   { background: rgba(255,165,2,0.1);  border: 1px solid rgba(255,165,2,0.3);  border-radius: 10px; padding: 10px 14px; color: #ffa502; font-weight: 700; font-size: 13px; }
.badge-active  { background: rgba(46,213,115,0.1); border: 1px solid rgba(46,213,115,0.3); border-radius: 10px; padding: 10px 14px; color: #2ed573; font-weight: 700; font-size: 13px; }
.badge-expired { background: rgba(255,71,87,0.1);  border: 1px solid rgba(255,71,87,0.3);  border-radius: 10px; padding: 10px 14px; color: #ff4757; font-weight: 700; font-size: 13px; }

.section-title { font-size: 18px; font-weight: 800; color: #ffffff; margin: 24px 0 16px; display: flex; align-items: center; gap: 10px; }
.section-title span { font-size: 14px; font-weight: 500; color: #555577; }
</style>
""", unsafe_allow_html=True)

# Manage app + Badge yashirish + UID localStorage
st.markdown("""<script>
(function(){
  // 1. Manage app va belgilarni yashirish
  var css='[data-testid="stToolbar"],[data-testid="stStatusWidget"],[data-testid="stBottom"],'
    +'[data-testid="manage-app-button"],[title="Manage app"],[class*="viewerBadge"],'
    +'[class*="StatusWidget"],[class*="ActionButton"],[class*="action-button"],'
    +'[data-testid="stActionButton"],a[href*="streamlit.io"],a[href*="share.streamlit"],'
    +'#stDecoration,footer{display:none!important;visibility:hidden!important;}';
  var el=document.createElement('style');
  el.appendChild(document.createTextNode(css));
  document.head.appendChild(el);

  // 2. MutationObserver bilan dinamik yashirish
  var mo=new MutationObserver(function(){
    document.querySelectorAll(
      '[data-testid="manage-app-button"],[title="Manage app"],'
      +'[class*="ActionButton"],[class*="action-button"],'
      +'a[href*="streamlit.io"],a[href*="share.streamlit"]'
    ).forEach(function(n){
      n.style.cssText='display:none!important;visibility:hidden!important;';
    });
  });
  mo.observe(document.body,{childList:true,subtree:true});

  // 3. UID localStorage da saqlash — yangi oynada ham bir xil UID
  try {
    var urlParams = new URLSearchParams(window.location.search);
    var urlUid = urlParams.get('uid');
    var storedUid = localStorage.getItem('viral777_uid');

    if (urlUid && urlUid.length > 10) {
      // URL da UID bor — saqlash
      localStorage.setItem('viral777_uid', urlUid);
    } else if (storedUid && storedUid.length > 10 && !urlUid) {
      // URL da UID yo'q, lekin localStorage da bor — URLga qo'shish
      var newUrl = window.location.href + 
        (window.location.search ? '&' : '?') + 'uid=' + storedUid;
      window.location.replace(newUrl);
    }
  } catch(e) {}
})();
</script>""", unsafe_allow_html=True)

# ══════════════════════════════════════════
# DATABASE
# ══════════════════════════════════════════
def load_db():
    if os.path.exists(DB_FILE):
        try:
            with open(DB_FILE) as f: return json.load(f)
        except: pass
    return {}

def save_db(data):
    with open(DB_FILE, "w") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

def get_user(uid):
    db = load_db()
    if uid not in db:
        db[uid] = {"trial_used":0,"subscribed":False,"sub_until":None,"orders":[]}
        save_db(db)
    return db[uid]

def save_user(uid, data):
    db = load_db(); db[uid] = data; save_db(db)

def is_subscribed(uid):
    u = get_user(uid)
    if not u.get("subscribed"): return False
    if u.get("sub_until"):
        if datetime.now() > datetime.fromisoformat(u["sub_until"]):
            u["subscribed"]=False; u["sub_until"]=None; save_user(uid,u); return False
    return True

def get_trial(uid):
    return max(0, FREE_TRIAL_LIMIT - get_user(uid).get("trial_used",0))

def use_trial(uid):
    u = get_user(uid); u["trial_used"] = u.get("trial_used",0)+1; save_user(uid,u)

def activate_sub(uid, order_id):
    u = get_user(uid)
    u["subscribed"] = True
    u["sub_until"]  = (datetime.now()+timedelta(days=SUBSCRIPTION_DAYS)).isoformat()
    u.setdefault("orders",[]).append({"id":order_id,"date":datetime.now().isoformat()})
    save_user(uid,u)

def gen_code():
    db = load_db(); codes = db.get("activation_codes",{})
    for _ in range(200):
        c = ''.join(random.choices(string.ascii_uppercase+string.digits, k=6))
        if c not in codes: return c
    return uuid.uuid4().hex[:6].upper()

def save_code(code, tg_id, order_id, note="", days=30):
    db = load_db()
    db.setdefault("activation_codes",{})[code] = {
        "telegram_id":str(tg_id), "order_id":order_id, "note":note,
        "created":datetime.now().isoformat(),
        "expires":(datetime.now()+timedelta(days=days)).isoformat(), "used":False
    }
    save_db(db)

def activate_by_code(code, uid):
    db = load_db(); codes = db.get("activation_codes",{})
    code = code.strip().upper()
    if code not in codes: return False,"❌ Kod topilmadi."
    c = codes[code]
    if c.get("used"):    return False,"❌ Bu kod allaqachon ishlatilgan."
    if datetime.now() > datetime.fromisoformat(c["expires"]): return False,"❌ Kod muddati tugagan."
    activate_sub(uid, code)
    codes[code].update({"used":True,"used_by":uid,"used_at":datetime.now().isoformat()})
    db["activation_codes"] = codes; save_db(db)
    return True,"✅ Obuna muvaffaqiyatli faollashdi!"

# ══════════════════════════════════════════
# EXCEL YARATISH FUNKSIYASI
# ══════════════════════════════════════════
def create_excel(df, topic_name):
    """Chiroyli formatlangan Excel fayl yaratish"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Viral Tahlil"

    # Ranglar
    PURPLE = "6C63FF"
    DARK   = "0F0F1A"
    HEADER = "1A1A2E"
    RED    = "FF4757"
    ORANGE = "FFA502"
    BLUE   = "6C63FF"
    GREEN  = "2ED573"
    WHITE  = "FFFFFF"
    GRAY   = "AAAACC"

    # ── Sarlavha qatori (1-qator) ──
    ws.merge_cells("A1:I1")
    ws["A1"] = f"🔥 Viral 777 — {topic_name} Nishasi Tahlili"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=PURPLE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # ── Ustun sarlavhalari (2-qator) ──
    headers = ["#", "Sarlavha", "Kanal", "⚡ Score", "👁 Ko'rishlar",
               "👥 Obunachi", "👍 Layklar", "💬 Engage%", "🔗 Havola"]
    col_widths = [5, 55, 22, 10, 14, 13, 12, 11, 50]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(name="Arial", bold=True, size=10, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 24

    # ── Ma'lumotlar ──
    thin = Side(style="thin", color="2A2A4A")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, (_, row) in enumerate(df.iterrows(), 1):
        r = i + 2
        score = row["outlier"]
        engage = round((row["likes"]+row["comments"])/max(row["views"],1)*100, 2)

        # Score rangi
        if score >= 100:   sc_color = RED
        elif score >= 50:  sc_color = ORANGE
        elif score >= 20:  sc_color = BLUE
        else:              sc_color = GREEN

        row_bg = "0A0A14" if i % 2 == 0 else "0F0F1A"

        data = [
            i,
            row["title"],
            row["channel"],
            f"{score}x",
            int(row["views"]),
            int(row["subs"]),
            int(row["likes"]),
            f"{engage}%",
            row["url"],
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = Font(name="Arial", size=9,
                             color=sc_color if col == 4 else (WHITE if col != 9 else "6C63FF"))
            cell.fill = PatternFill("solid", fgColor=row_bg)
            cell.alignment = Alignment(
                horizontal="center" if col in [1,4,5,6,7,8] else "left",
                vertical="center", wrap_text=(col == 2)
            )
            cell.border = border
            if col == 4:
                cell.font = Font(name="Arial", bold=True, size=11, color=sc_color)
            if col == 5:
                cell.number_format = "#,##0"
            if col == 6:
                cell.number_format = "#,##0"
            if col == 7:
                cell.number_format = "#,##0"
        ws.row_dimensions[r].height = 20

    # ── Freeze panes ──
    ws.freeze_panes = "A3"

    # ── 2-sheet: Statistika ──
    ws2 = wb.create_sheet("Statistika")
    ws2["A1"] = "Ko'rsatkich"
    ws2["B1"] = "Qiymat"
    ws2["A1"].font = Font(bold=True, color=WHITE, name="Arial")
    ws2["A1"].fill = PatternFill("solid", fgColor=PURPLE)
    ws2["B1"].font = Font(bold=True, color=WHITE, name="Arial")
    ws2["B1"].fill = PatternFill("solid", fgColor=PURPLE)
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 18

    stats = [
        ("Mavzu",            topic_name),
        ("Jami videolar",    len(df)),
        ("O'rtacha Score",   f"{round(df['outlier'].mean(),1)}x"),
        ("Eng yuqori Score", f"{df['outlier'].max()}x"),
        ("Jami Ko'rishlar",  int(df['views'].sum())),
        ("O'rtacha Ko'rish", int(df['views'].mean())),
        ("Top Kanal",        df.iloc[0]['channel'] if len(df)>0 else "-"),
    ]
    for r, (k, v) in enumerate(stats, 2):
        ws2.cell(row=r, column=1, value=k).font = Font(name="Arial", size=10)
        ws2.cell(row=r, column=2, value=v).font = Font(name="Arial", bold=True, size=10, color=PURPLE)

    # Bytes ga o'girish
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ══════════════════════════════════════════
# TELEGRAM LOGIN TIZIMI
# ══════════════════════════════════════════
TG_LOGIN_BOT_TOKEN = "8356461690:AAHlgC3e3SQCmCPKM7H7wO3YFhSD4HSNGlM"

def gen_login_code():
    """6 xonali login kodi yaratish"""
    return ''.join(random.choices(string.digits, k=6))

def save_login_code(code, tg_id, tg_name=""):
    """Login kodni DB ga saqlash (5 daqiqa amal qiladi)"""
    db = load_db()
    db.setdefault("login_codes", {})[code] = {
        "tg_id":   str(tg_id),
        "tg_name": tg_name,
        "created": datetime.now().isoformat(),
        "expires": (datetime.now() + timedelta(minutes=5)).isoformat(),
        "used":    False,
    }
    save_db(db)

def verify_login_code(code):
    """Login kodni tekshirish — (ok, tg_id, tg_name)"""
    db = load_db()
    codes = db.get("login_codes", {})
    code = code.strip()
    if code not in codes:
        return False, None, None
    c = codes[code]
    if c.get("used"):
        return False, None, None
    if datetime.now() > datetime.fromisoformat(c["expires"]):
        return False, None, None
    # Ishlatilgan deb belgilash
    codes[code]["used"] = True
    db["login_codes"] = codes
    save_db(db)
    return True, c["tg_id"], c.get("tg_name","")

def get_tg_uid(tg_id):
    """Telegram ID dan UID yaratish"""
    return f"tg_{tg_id}"

def send_login_code(tg_id, code):
    """Botdan login kodi yuborish"""
    try:
        import urllib.request
        msg = ("🔐 *Viral 777 — Kirish Kodi*\n\n"
               f"`{code}`\n\n"
               "⏰ Kod *5 daqiqa* amal qiladi.\n\n"
               "_Agar siz kirmagan bo\'lsangiz, e\'tibor bermang._")
        url = f"https://api.telegram.org/bot{TG_LOGIN_BOT_TOKEN}/sendMessage"
        data = json.dumps({"chat_id": tg_id, "text": msg, "parse_mode": "Markdown"}).encode()
        req = urllib.request.Request(url, data=data,
                                     headers={"Content-Type":"application/json"})
        urllib.request.urlopen(req, timeout=5)
        return True
    except Exception as e:
        return False

def setup_login_bot():
    """Login bot /start handler"""
    if not BOT_AVAILABLE or not TG_LOGIN_BOT_TOKEN: return
    try:
        login_bot = telebot.TeleBot(TG_LOGIN_BOT_TOKEN, threaded=False)

        @login_bot.message_handler(commands=['start', 'login', 'kirish'])
        def handle_login(msg):
            tg_id   = msg.from_user.id
            tg_name = msg.from_user.first_name or "Foydalanuvchi"
            code    = gen_login_code()
            save_login_code(code, tg_id, tg_name)
            login_bot.send_message(
                tg_id,
                f"👋 Salom, *{tg_name}*!\n\n"
                f"🔐 *Viral 777 kirish kodi:*\n\n"
                f"`{code}`\n\n"
                f"⏰ Kod *5 daqiqa* amal qiladi.\n"
                f"📌 Saytga qaytib kodni kiriting.",
                parse_mode="Markdown"
            )

        @login_bot.message_handler(commands=['obuna', 'start_pay'])
        def handle_pay(msg):
            from_id = msg.from_user.id
            login_bot.send_message(
                from_id,
                f"💰 *Obuna sotib olish:*\n\n"
                f"📅 30 kun — {SUBSCRIPTION_PRICE:,} so'm\n\n"
                f"To'lov qilish uchun admin bilan bog'laning.",
                parse_mode="Markdown"
            )

        def poll_login():
            while True:
                try: login_bot.infinity_polling(timeout=15, long_polling_timeout=10)
                except: time.sleep(5)

        threading.Thread(target=poll_login, daemon=True).start()
    except: pass

# ══════════════════════════════════════════
# PAYMENT URLS
# ══════════════════════════════════════════
def payme_url(oid):
    p = f"m={PAYME_MERCHANT};ac.order_id={oid};a={SUBSCRIPTION_PRICE*100}"
    return f"https://checkout.paycom.uz/{base64.b64encode(p.encode()).decode()}"

def click_url(oid):
    return (f"https://my.click.uz/services/pay?service_id={CLICK_SERVICE}"
            f"&merchant_id={CLICK_MERCHANT}&amount={SUBSCRIPTION_PRICE}&transaction_param={oid}")

# ══════════════════════════════════════════
# TELEGRAM BOT
# ══════════════════════════════════════════
def start_bot():
    if not BOT_AVAILABLE or not BOT_TOKEN: return
    bot = telebot.TeleBot(BOT_TOKEN, threaded=False)

    @bot.message_handler(commands=['start'])
    def h_start(msg):
        oid = f"V777_{uuid.uuid4().hex[:10].upper()}"
        mk = telebot.types.InlineKeyboardMarkup(row_width=1)
        mk.add(
            telebot.types.InlineKeyboardButton("💳 Payme orqali to'lash",  url=payme_url(oid)),
            telebot.types.InlineKeyboardButton("⚡ Click orqali to'lash",   url=click_url(oid)),
            telebot.types.InlineKeyboardButton("✅ To'lov qildim — kod olish", callback_data=f"paid:{oid}"),
        )
        bot.send_message(msg.chat.id,
            f"👋 *Viral 777 Analytics*ga xush kelibsiz!\n\n"
            f"🔥 YouTube viral tahlil platformasi\n\n"
            f"💰 *Oylik obuna:* `{SUBSCRIPTION_PRICE:,} so'm`\n"
            f"📅 30 kun · ♾️ Cheksiz qidiruv · 📊 Trend tahlili\n\n"
            f"To'lovni amalga oshirib, *\"To'lov qildim\"* tugmasini bosing:",
            parse_mode="Markdown", reply_markup=mk)

    @bot.callback_query_handler(func=lambda c: c.data.startswith("paid:"))
    def h_paid(call):
        oid = call.data.split(":",1)[1]
        tg_id = call.from_user.id
        bot.answer_callback_query(call.id, "⏳ Tekshirilmoqda...")
        code = gen_code()
        save_code(code, tg_id, oid)
        bot.send_message(tg_id,
            f"✅ *To'lov qabul qilindi!*\n\n🔑 Aktivatsiya kodingiz:\n\n"
            f"```\n{code}\n```\n\n📌 Saytga qayting va ushbu kodni kiriting.\n"
            f"⏰ Kod *24 soat* amal qiladi.", parse_mode="Markdown")
        if ADMIN_CHAT_ID:
            try:
                bot.send_message(int(ADMIN_CHAT_ID),
                    f"💰 *Yangi to'lov!*\n👤 {call.from_user.first_name} (`{tg_id}`)\n"
                    f"📋 `{oid}`\n🔑 `{code}`\n💵 {SUBSCRIPTION_PRICE:,} so'm\n"
                    f"🕐 {datetime.now().strftime('%d.%m.%Y %H:%M')}", parse_mode="Markdown")
            except: pass

    @bot.message_handler(commands=['admin'])
    def h_admin(msg):
        if str(msg.chat.id)!=str(ADMIN_CHAT_ID): return
        db=load_db(); u={k:v for k,v in db.items() if k!="activation_codes"}
        c=db.get("activation_codes",{})
        bot.send_message(msg.chat.id,
            f"📊 *Statistika*\n\n👥 Foydalanuvchi: `{len(u)}`\n"
            f"✅ Faol obuna: `{sum(1 for x in u.values() if x.get('subscribed'))}`\n"
            f"🔑 Ishlatilgan: `{sum(1 for x in c.values() if x.get('used'))}`\n"
            f"⏳ Kutilayotgan: `{sum(1 for x in c.values() if not x.get('used'))}`",
            parse_mode="Markdown")

    @bot.message_handler(commands=['kod'])
    def h_kod(msg):
        if str(msg.chat.id)!=str(ADMIN_CHAT_ID): return
        parts=msg.text.split()
        if len(parts)<2: bot.send_message(msg.chat.id,"Format: /kod TELEGRAM\\_ID",parse_mode="Markdown"); return
        code=gen_code(); save_code(code, parts[1], f"MANUAL_{uuid.uuid4().hex[:6]}", note="manual")
        try:
            bot.send_message(int(parts[1]),
                f"✅ *Aktivatsiya kodingiz:*\n\n```\n{code}\n```\n\nSaytga kiring va kodni kiriting.",
                parse_mode="Markdown")
            bot.send_message(msg.chat.id, f"✅ Yuborildi: `{code}`", parse_mode="Markdown")
        except Exception as e:
            bot.send_message(msg.chat.id, f"❌ {e}")

    def poll():
        while True:
            try: bot.infinity_polling(timeout=20, long_polling_timeout=15)
            except: time.sleep(5)

    threading.Thread(target=poll, daemon=True).start()

if "bot_started" not in st.session_state:
    st.session_state.bot_started = True
    threading.Thread(target=start_bot, daemon=True).start()
    # Login bot (alohida token)
    if TG_LOGIN_BOT_TOKEN != cfg("BOT_TOKEN",""):
        threading.Thread(target=setup_login_bot, daemon=True).start()

# ══════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════
def fmt(n):
    if n>=1_000_000_000: return f"{round(n/1_000_000_000,1)}B"
    if n>=1_000_000:     return f"{round(n/1_000_000,1)}M"
    if n>=1_000:         return f"{round(n/1_000,1)}K"
    return str(n)

def uzb_date(iso):
    M={1:"Yan",2:"Fev",3:"Mar",4:"Apr",5:"May",6:"Iyn",
       7:"Iyl",8:"Avg",9:"Sen",10:"Okt",11:"Noy",12:"Dek"}
    try:
        dt=datetime.strptime(re.sub(r'\.\d+Z','Z',iso),'%Y-%m-%dT%H:%M:%SZ')
        now=datetime.utcnow(); diff=now-dt
        if diff.days==0:    return "Bugun"
        if diff.days==1:    return "Kecha"
        if diff.days<7:     return f"{diff.days} kun oldin"
        if diff.days<30:    return f"{diff.days//7} hafta oldin"
        if diff.days<365:   return f"{diff.days//30} oy oldin"
        return f"{dt.day} {M[dt.month]} {dt.year}"
    except: return iso[:10]

def get_uid():
    """UID — URL da saqlanadi, yangi oynada ham bir xil bo'lishi uchun
    foydalanuvchi UID ni URL orqali olib yuradi."""
    if "uid" in st.session_state and st.session_state.uid:
        return st.session_state.uid
    p = st.query_params.get("uid","")
    if p and len(p)>=10:
        st.session_state.uid=p; return p
    new = "u_"+uuid.uuid4().hex
    st.session_state.uid=new; st.query_params["uid"]=new; return new

def viral_badge(score):
    if score>=100: return '<span class="vc-badge fire">🔥 Mega Viral</span>'
    if score>=50:  return '<span class="vc-badge fire">🚀 Viral</span>'
    if score>=20:  return '<span class="vc-badge hot">⚡ Trendda</span>'
    return '<span class="vc-badge ok">✅ Yaxshi</span>'

NICHES = [
    "Survival","Cooking","Finance","Gaming","Science","Travel",
    "Fitness","Tech","Psychology","History","Cars","Football",
    "DIY","Comedy","Documentary","Music","Fashion","Business",
]

REGIONS = {"🇺🇸 US":"US","🇬🇧 GB":"GB","🇺🇿 UZ":"UZ",
           "🇷🇺 RU":"RU","🇹🇷 TR":"TR","🇩🇪 DE":"DE","🇯🇵 JP":"JP"}

# ══════════════════════════════════════════
# SESSION
# ══════════════════════════════════════════
for k,v in [("authenticated",False),("results",None),
             ("history",[]),("last_topic",""),("search_done",False),
             ("dark_mode",True),("current_topic","Survival"),
             ("topic_key_ver",0),("do_search",False),
             ("tg_logged_in",False),("tg_id",None),("tg_name","")]:
    if k not in st.session_state: st.session_state[k]=v

# History fayldan tiklash
if not st.session_state.history:
    try:
        _uid_tmp = st.query_params.get("uid","")
        if _uid_tmp:
            hist_file = f"/tmp/v777_hist_{_uid_tmp[:12]}.json"
            if os.path.exists(hist_file):
                with open(hist_file) as hf:
                    st.session_state.history = json.load(hf)
    except: pass

# Admin login
_admin_secret = "v777admin2024"
if st.query_params.get("au","") == _admin_secret:
    st.session_state.authenticated = True

# Agar Telegram login qilingan bo'lsa — TG ID asosida UID
if st.session_state.get("tg_logged_in") and st.session_state.get("tg_id"):
    uid = get_tg_uid(st.session_state.tg_id)
    st.session_state.uid = uid
else:
    uid = get_uid()

# Share URL
try:
    _base_url = st.secrets["general"].get("APP_URL","https://my-youtube-trends-nk9rk2csy8hjj6dvhfbgku.streamlit.app")
except:
    _base_url = "https://my-youtube-trends-nk9rk2csy8hjj6dvhfbgku.streamlit.app"
current_url = _base_url

# ══════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════
with st.sidebar:
    apply_theme()

    # Theme + Share
    t1, t2 = st.columns([1, 1])
    dark_icon = "🌙 Dark" if st.session_state.dark_mode else "☀️ Light"
    if t1.button(dark_icon, key="theme_btn", use_container_width=True):
        st.session_state.dark_mode = not st.session_state.dark_mode
        st.rerun()
    _app_url = "https://my-youtube-trends-nk9rk2csy8hjj6dvhfbgku.streamlit.app/"
    tg_share = f"https://t.me/share/url?url={_app_url}&text=🔥%20Viral%20777%20-%20YouTube%20Trend%20Analytics%20platformasi"
    t2.markdown(
        f"<a href='{tg_share}' target='_blank' style='display:block;text-align:center;"
        f"background:#229ED9;color:#fff;border-radius:10px;padding:6px 4px;"
        f"font-weight:700;font-size:13px;text-decoration:none;'>📤 Share</a>",
        unsafe_allow_html=True
    )

    # Logo
    st.markdown("""
    <div class="vs-header">
        <div>
            <div class="vs-logo">🔥 Viral 777</div>
            <div class="vs-tagline">YouTube Analytics Platform</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Auth — faqat ?admin=1 yoki allaqachon kirgan bo'lsa
    _show_admin = (st.query_params.get("admin","")=="1" or st.session_state.authenticated)
    if _show_admin and not st.session_state.authenticated:
        with st.expander("🔐 Admin Kirish", expanded=True):
            u_in = st.text_input("Login", key="login_u")
            p_in = st.text_input("Parol", type="password", key="login_p")
            if st.button("Kirish", key="login_btn"):
                if ADMIN_DB.get(u_in)==p_in:
                    st.session_state.authenticated=True
                    st.query_params["au"]="v777admin2024"
                    st.rerun()
                else: st.error("❌ Notogri!")
    elif st.session_state.authenticated:
        col1,col2 = st.columns([2,1])
        col1.success("✅ Admin")
        if col2.button("Chiqish"):
            st.session_state.authenticated=False
            for k in ["au","admin"]:
                if k in st.query_params: del st.query_params[k]
            st.rerun()

    st.divider()

    # ══ TELEGRAM LOGIN ══
    if not st.session_state.authenticated:
        if st.session_state.get("tg_logged_in"):
            # Kirgan holat
            tg_nm = st.session_state.get("tg_name","Foydalanuvchi")
            st.markdown(
                f"<div style='background:rgba(34,158,217,0.1);border:1px solid rgba(34,158,217,0.3);"
                f"border-radius:10px;padding:10px 14px;margin-bottom:8px;'>"
                f"<span style='color:#229ED9;font-weight:700;'>✈️ {tg_nm}</span>"
                f"<span style='color:#555577;font-size:11px;'> bilan kirildi</span></div>",
                unsafe_allow_html=True
            )
            if st.button("🚪 Chiqish", key="tg_logout", use_container_width=True):
                st.session_state.tg_logged_in = False
                st.session_state.tg_id   = None
                st.session_state.tg_name = ""
                st.rerun()
        else:
            # Login forma
            _tg_bot_username = "viral777_bot"  # @BotFather da olgan username
            st.markdown(
                f"<a href='https://t.me/{_tg_bot_username}?start=login' target='_blank' "
                f"style='display:block;text-align:center;background:#229ED9;color:#fff;"
                f"border-radius:10px;padding:10px;font-weight:700;font-size:13px;"
                f"text-decoration:none;margin-bottom:8px;'>✈️ Telegram orqali kirish</a>",
                unsafe_allow_html=True
            )
            _lc = st.text_input("Bot bergan 6 xonali kodni kiriting:",
                                 max_chars=6, placeholder="123456",
                                 key="tg_login_code")
            if st.button("✅ Kodni Tasdiqlash", key="tg_code_btn", use_container_width=True):
                if len(_lc.strip()) == 6:
                    ok, tg_id, tg_nm = verify_login_code(_lc.strip())
                    if ok:
                        st.session_state.tg_logged_in = True
                        st.session_state.tg_id   = tg_id
                        st.session_state.tg_name = tg_nm
                        # UID ni TG ID ga bog'laymiz
                        st.session_state.uid = get_tg_uid(tg_id)
                        st.success(f"✅ Xush kelibsiz, {tg_nm}!")
                        time.sleep(0.5)
                        st.rerun()
                    else:
                        st.error("❌ Kod noto'g'ri yoki muddati tugagan!")
                else:
                    st.warning("6 xonali kod kiriting")

    st.divider()

    # Subscription status
    if not st.session_state.authenticated:
        if is_subscribed(uid):
            ud=get_user(uid)
            until=datetime.fromisoformat(ud["sub_until"]).strftime("%d.%m.%Y")
            st.markdown(f"<div class='badge-active'>✅ PRO · {until} gacha</div>",
                        unsafe_allow_html=True)
        else:
            rem=get_trial(uid)
            if rem>0:
                st.markdown(f"<div class='badge-trial'>🎁 Sinov: {rem}/{FREE_TRIAL_LIMIT}</div>",
                            unsafe_allow_html=True)
            else:
                st.markdown("<div class='badge-expired'>🔒 Obuna kerak</div>",
                            unsafe_allow_html=True)

    st.divider()

    # API Key
    saved_key = st.query_params.get("apikey","")
    with st.expander("🔑 YouTube API Key" + (" ✅" if saved_key else " ❗"), expanded=not bool(saved_key)):
        st.markdown("""
**API Key olish:**

1. 🌐 [console.cloud.google.com](https://console.cloud.google.com)
2. ➕ New Project → Create
3. APIs & Services → Library → **YouTube Data API v3** → Enable
4. Credentials → + Create Credentials → **API Key**
5. Kalitni nusxalab quyidagi maydonga joylashtiring ✅

⚠️ *Bepul: kuniga 10,000 so'rov*
        """)
    api_key = st.text_input("API Key:", value=saved_key, type="password",
                             placeholder="AIzaSy...", label_visibility="collapsed")
    if api_key and api_key != saved_key:
        st.query_params["apikey"] = api_key

    st.markdown("**🔍 Mavzu / Nisha**")

    # Niche quick select (chap panel)
    niche_cols = st.columns(3)
    niche_list = NICHES[:9]
    for i in range(len(niche_list)):
        niche = niche_list[i]
        is_active = st.session_state.get("current_topic","") == niche
        label = f"✓ {niche}" if is_active else niche
        if niche_cols[i%3].button(label, key=f"nch_{i}", use_container_width=True):
            st.session_state["current_topic"] = niche
            st.session_state["topic_key_ver"] = st.session_state.get("topic_key_ver",0)+1
            st.rerun()

    topic = st.text_input(
        "Yoki qidiruv matnini kiriting:",
        value=st.session_state["current_topic"],
        placeholder="Masalan: Survival, Finance...",
        key=f"topic_input_{st.session_state['topic_key_ver']}"
    )
    if topic != st.session_state["current_topic"]:
        st.session_state["current_topic"] = topic

    region_label = st.selectbox("🌍 Bozor:", list(REGIONS.keys()))
    region_code  = REGIONS[region_label]

    days_sel  = st.select_slider("📅 Davr:", options=[1,7,14,30,60,90,180,365],
                                  value=30, format_func=lambda x: f"{x} kun")
    min_outl  = st.slider("⚡ Min Outlier Score:", 1, 200, 10,
                           help="Viral koeffitsient — video ko'rishlari / obunachilarga nisbat")
    min_views = st.select_slider("👁 Min Ko'rishlar:",
                                  options=[0,1000,10000,50000,100000,500000,1000000],
                                  value=0, format_func=lambda x: fmt(x) if x>0 else "Hammasi")
    max_res   = st.select_slider("📦 Natijalar soni:", options=[10,25,50], value=25)

    # Qidirish uchun: admin YOKI tg_logged_in YOKI sinov qolgan
    can_search = (
        st.session_state.authenticated or
        st.session_state.get("tg_logged_in", False) and (is_subscribed(uid) or get_trial(uid)>0)
    )
    # Telegram login bo'lmagan bo'lsa ham 1 ta bepul sinov
    if not st.session_state.get("tg_logged_in") and not st.session_state.authenticated:
        can_search = get_trial(uid) > 0
    search_btn = st.button("🚀 TAHLILNI BOSHLASH", disabled=not can_search,
                            use_container_width=True)
    if not can_search:
        st.markdown("<div style='text-align:center;margin-top:8px;font-size:12px;color:#555577;'>"
                    "Obuna sotib oling ↓</div>", unsafe_allow_html=True)

# ══════════════════════════════════════════
# MAIN TABS
# ══════════════════════════════════════════
TAB_TREND, TAB_CARDS, TAB_TABLE, TAB_CHART, TAB_HISTORY, TAB_GUIDE, TAB_ADMIN = st.tabs([
    "🔥 Trend Tahlili",
    "🎬 Video Kartochkalar",
    "📊 Jadval",
    "📈 Grafiklar",
    "🕐 Tarix",
    "📖 Qo\'llanma",
    "🛡️ Admin",
])

# ══════════════════════════════════════════
# SUBSCRIPTION BLOCK
# ══════════════════════════════════════════
def show_sub_block(tab_id="main"):
    oid = st.session_state.get("pending_order") or gen_code()
    st.session_state.pending_order = oid
    st.markdown(f"""
    <div class='sub-box'>
        <div style='font-size:48px;margin-bottom:12px;'>🔒</div>
        <h2 style='color:#fff;font-size:24px;margin-bottom:8px;'>Sinov muddati tugadi</h2>
        <p style='color:#888899;font-size:14px;margin-bottom:24px;'>
            Davom etish uchun <b>Pro obuna</b> sotib oling.<br>
            To'lovdan so'ng Telegram bot aktivatsiya kodini <b>darhol</b> yuboradi.
        </p>
        <div style='font-size:52px;font-weight:900;background:linear-gradient(135deg,#6c63ff,#ff6b6b);
                    -webkit-background-clip:text;-webkit-text-fill-color:transparent;margin:16px 0;'>
            {SUBSCRIPTION_PRICE:,} so'm
        </div>
        <p style='color:#444466;font-size:13px;margin-bottom:28px;'>
            📅 30 kun &nbsp;·&nbsp; ♾️ Cheksiz qidiruv &nbsp;·&nbsp; 📊 Trend tahlili
        </p>
        <a href='{TELEGRAM_BOT_LINK}' target='_blank' class='tg-pay-btn'>
            ✈️ Telegram orqali to'lash va kod olish
        </a>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("""
    <div class='act-box'>
        <div style='font-size:28px;margin-bottom:8px;'>🔑</div>
        <h4 style='color:#fff;margin-bottom:4px;'>Aktivatsiya kodini kiriting</h4>
        <p style='color:#555577;font-size:13px;'>Telegram botdan olgan 6 ta belgili kod</p>
    </div>
    """, unsafe_allow_html=True)
    c1,c2,c3 = st.columns([1,2,1])
    with c2:
        code = st.text_input("", placeholder="4X9K2M", max_chars=6,
                              key=f"code_inp_{tab_id}",
                              label_visibility="collapsed").strip().upper()
        if st.button("🔓 KODNI TASDIQLASH", key=f"code_btn_{tab_id}", use_container_width=True):
            if len(code)<6: st.error("❌ Kod 6 ta belgidan iborat!")
            else:
                ok,msg = activate_by_code(code, uid)
                if ok: st.success(msg); st.balloons(); time.sleep(1); st.rerun()
                else:  st.error(msg)

# ══════════════════════════════════════════
# SEARCH LOGIC
# ══════════════════════════════════════════
# do_search — welcome screen nishasi bosilganda True bo'ladi
_trigger = search_btn or st.session_state.get("do_search", False)
if st.session_state.get("do_search", False):
    st.session_state["do_search"] = False

if _trigger:
    topic = st.session_state.get("current_topic", "Survival")
    current_key = st.query_params.get("apikey","") or api_key
    if not current_key:
        st.error("‼️ YouTube API Key kiritilmagan!")
    else:
        with st.spinner("🔍 YouTube ma'lumotlari tahlil qilinmoqda..."):
            try:
                yt = googleapiclient.discovery.build("youtube","v3",developerKey=current_key)
                pub_after = (datetime.utcnow()-timedelta(days=days_sel)).isoformat()+"Z"
                res = yt.search().list(
                    q=topic, part="snippet", type="video",
                    maxResults=max_res, order="viewCount",
                    publishedAfter=pub_after, regionCode=region_code
                ).execute()

                results = []
                vid_ids = [item['id']['videoId'] for item in res.get('items',[])]
                ch_ids  = [item['snippet']['channelId'] for item in res.get('items',[])]

                vi_batch = yt.videos().list(part="statistics,snippet", id=",".join(vid_ids)).execute()
                vi_map = {i['id']:i for i in vi_batch.get('items',[])}

                unique_chs = list(set(ch_ids))
                ci_batch = yt.channels().list(part="statistics,snippet", id=",".join(unique_chs)).execute()
                ci_map = {i['id']:i for i in ci_batch.get('items',[])}

                for item in res.get('items',[]):
                    vid = item['id']['videoId']
                    cid = item['snippet']['channelId']
                    vi  = vi_map.get(vid, {})
                    ci  = ci_map.get(cid, {})
                    if not vi or not ci: continue
                    views    = int(vi.get('statistics',{}).get('viewCount',0))
                    subs     = int(ci.get('statistics',{}).get('subscriberCount',1))
                    likes    = int(vi.get('statistics',{}).get('likeCount',0))
                    comments = int(vi.get('statistics',{}).get('commentCount',0))
                    outl     = round(views/(subs if subs>500 else 500),1)
                    thumbs   = vi.get('snippet',{}).get('thumbnails',{})
                    thumb    = (thumbs.get('maxres') or thumbs.get('high') or
                                thumbs.get('medium') or thumbs.get('default',{})).get('url','')
                    # Teglar
                    tags = vi.get('snippet',{}).get('tags',[]) or []
                    tags_str = ", ".join(tags[:8]) if tags else "—"
                    # Taxminiy daromad ($2.5 RPM)
                    est_monthly  = round(views / 1000 * 2.5, 1)
                    is_monetized = subs >= 1000

                    if outl>=min_outl and views>=min_views:
                        results.append({
                            "id":vid, "thumbnail":thumb,
                            "title":vi.get('snippet',{}).get('title',''),
                            "channel":item['snippet']['channelTitle'],
                            "channel_id":cid, "views":views, "subs":subs,
                            "likes":likes, "comments":comments, "outlier":outl,
                            "published":vi.get('snippet',{}).get('publishedAt',''),
                            "url":f"https://www.youtube.com/watch?v={vid}",
                            "ch_url":f"https://www.youtube.com/channel/{cid}",
                            "tags":         tags_str,
                            "est_income":   est_monthly,
                            "is_monetized": is_monetized,
                        })

                results.sort(key=lambda x: x['outlier'], reverse=True)
                st.session_state.results = results
                st.session_state.last_topic = topic
                st.session_state.search_done = True

                now_uz = datetime.utcnow() + timedelta(hours=5)
                new_entry = {
                    "Vaqt": now_uz.strftime("%H:%M"), "Sana": now_uz.strftime("%d.%m"),
                    "Mavzu": topic, "Bozor": region_label,
                    "Davr": f"{days_sel} kun", "Topildi": len(results),
                }
                st.session_state.history.append(new_entry)
                try:
                    hist_file = f"/tmp/v777_hist_{uid[:12]}.json"
                    existing = []
                    if os.path.exists(hist_file):
                        with open(hist_file) as hf: existing = json.load(hf)
                    existing.append(new_entry)
                    existing = existing[-50:]
                    with open(hist_file, "w") as hf: json.dump(existing, hf)
                except: pass

                if not st.session_state.authenticated and not is_subscribed(uid):
                    use_trial(uid)
                    rem = get_trial(uid)
                    if rem>0: st.toast(f"🎁 Yana {rem} ta tekin qidiruv qoldi")
                    else:     st.toast("⚠️ Oxirgi sinov ishlatildi!")

            except Exception as e:
                st.error(f"⚠️ Xato: {e}")

# ══════════════════════════════════════════
# TAB 1: TREND
# ══════════════════════════════════════════
with TAB_TREND:
    need_sub = (not st.session_state.authenticated and
                not is_subscribed(uid) and get_trial(uid)==0)

    if need_sub:
        show_sub_block()
    elif not st.session_state.results:
        # Welcome screen
        st.markdown("""
        <div style='text-align:center;padding:60px 20px;'>
            <div style='font-size:64px;margin-bottom:16px;'>🔥</div>
            <h2 style='color:#fff;font-size:28px;margin-bottom:12px;'>
                YouTube Trend Tahlilchisi
            </h2>
            <p style='color:#555577;font-size:16px;max-width:500px;margin:0 auto 32px;'>
                Chap paneldan mavzu va parametrlarni tanlang,<br>
                so'ng <b>"Tahlilni boshlash"</b> tugmasini bosing.
            </p>
        </div>
        """, unsafe_allow_html=True)

        # ═══ 12 TA NISHA TUGMASI ═══
        # Bosilganda: chap inputga tushadi + qidiruv boshlanadi
        st.markdown("<div class='section-title'>⚡ Tez Qidiruv — Mashhur Nishalar</div>",
                    unsafe_allow_html=True)
        niche_emojis = {
            "Survival":"🏕️","Cooking":"🍳","Finance":"💰","Gaming":"🎮",
            "Science":"🔬","Travel":"✈️","Fitness":"💪","Tech":"💻",
            "Psychology":"🧠","History":"📜","Cars":"🚗","Football":"⚽"
        }
        nc = st.columns(6)
        for i, n in enumerate(NICHES[:12]):
            if nc[i%6].button(f"{niche_emojis.get(n,'🔥')} {n}", key=f"qn_{n}",
                               use_container_width=True):
                # 1. Chap paneldagi inputga yozish
                st.session_state["current_topic"] = n
                st.session_state["topic_key_ver"] = st.session_state.get("topic_key_ver",0)+1
                # 2. Qidiruvni trigger qilish
                st.session_state["do_search"] = True
                st.rerun()
    else:
        df = pd.DataFrame(st.session_state.results)
        if df.empty:
            st.info("Natija topilmadi. Parametrlarni o'zgartiring.")
        else:
            avg_outl    = round(df['outlier'].mean(),1)
            total_views = df['views'].sum()
            max_outl    = df['outlier'].max()
            top_channel = df.iloc[0]['channel']

            st.markdown(
                f"<div class='section-title'>📊 <b>{st.session_state.last_topic}</b> Nishasi — Umumiy Tahlil "
                f"<span>({len(df)} ta viral video topildi)</span></div>",
                unsafe_allow_html=True)

            k1,k2,k3,k4 = st.columns(4)
            k1.markdown(f"""<div class='stat-card purple'>
                <div class='stat-icon'>⚡</div>
                <div class='stat-label'>O'RTACHA VIRAL SCORE</div>
                <div class='stat-value'>{avg_outl}x</div>
                <div class='stat-sub'>Obunachilarga nisbatan</div>
            </div>""", unsafe_allow_html=True)
            k2.markdown(f"""<div class='stat-card red'>
                <div class='stat-icon'>👁</div>
                <div class='stat-label'>JAMI KO'RISHLAR</div>
                <div class='stat-value'>{fmt(total_views)}</div>
                <div class='stat-sub'>{days_sel} kunlik natija</div>
            </div>""", unsafe_allow_html=True)
            k3.markdown(f"""<div class='stat-card gold'>
                <div class='stat-icon'>🔥</div>
                <div class='stat-label'>ENG YUQORI SCORE</div>
                <div class='stat-value'>{max_outl}x</div>
                <div class='stat-sub'>Mega viral</div>
            </div>""", unsafe_allow_html=True)
            k4.markdown(f"""<div class='stat-card green'>
                <div class='stat-icon'>📺</div>
                <div class='stat-label'>TOP KANAL</div>
                <div class='stat-value' style='font-size:16px;'>{top_channel[:16]}...</div>
                <div class='stat-sub'>Eng viral kanal</div>
            </div>""", unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)

            col_gauge, col_bar = st.columns([1,2])
            with col_gauge:
                fig_gauge = go.Figure(go.Indicator(
                    mode="gauge+number", value=avg_outl,
                    title={'text':"Nisha Kuch Koeffitsienti",'font':{'color':'#aaaacc','size':12}},
                    number={'font':{'color':'#ffffff','size':40},'suffix':'x'},
                    gauge={
                        'axis':{'range':[0,max(200,avg_outl*1.5)],'tickcolor':'#333355','tickfont':{'color':'#666688'}},
                        'bar':{'color':'#6c63ff','thickness':0.3},
                        'bgcolor':'#0f0f1a','bordercolor':'#1e1e2e',
                        'steps':[
                            {'range':[0,10],'color':'#1a1a2e'},{'range':[10,50],'color':'#1e1e38'},
                            {'range':[50,100],'color':'#242448'},{'range':[100,max(200,avg_outl*1.5)],'color':'#2a2a58'},
                        ],
                        'threshold':{'line':{'color':'#ff4757','width':3},'value':50}
                    }
                ))
                fig_gauge.update_layout(paper_bgcolor='rgba(0,0,0,0)',plot_bgcolor='rgba(0,0,0,0)',
                    height=260,margin=dict(l=20,r=20,t=40,b=10),font=dict(family='Inter'))
                st.plotly_chart(fig_gauge, use_container_width=True)

            with col_bar:
                top10 = df.nlargest(10,'outlier')[['title','outlier','channel']].copy()
                top10['short_title'] = top10['title'].str[:35]+"..."
                fig_bar = px.bar(top10,x='outlier',y='short_title',orientation='h',color='outlier',
                    color_continuous_scale=['#2a2a58','#6c63ff','#ff4757'],
                    labels={'outlier':'Viral Score','short_title':''})
                fig_bar.update_layout(paper_bgcolor='rgba(0,0,0,0)',plot_bgcolor='rgba(0,0,0,0)',
                    height=260,margin=dict(l=0,r=20,t=20,b=10),
                    font=dict(color='#aaaacc',family='Inter',size=11),
                    xaxis=dict(gridcolor='#1e1e2e',color='#666688'),
                    yaxis=dict(gridcolor='#1e1e2e',color='#ccccdd'),
                    coloraxis_showscale=False)
                fig_bar.update_traces(marker_line_width=0)
                st.plotly_chart(fig_bar, use_container_width=True)

            st.markdown("<div class='section-title'>🏆 Eng Viral Videolar</div>", unsafe_allow_html=True)
            top3 = df.head(3)
            cols = st.columns(3)
            for i,(_, row) in enumerate(top3.iterrows()):
                with cols[i]:
                    badge_html = viral_badge(row['outlier'])
                    st.markdown(f"""
                    <div class='video-card'>
                        <img class='vc-thumbnail' src='{row['thumbnail']}'
                             onerror="this.src='https://via.placeholder.com/320x180/1a1a2e/6c63ff?text=No+Image'">
                        <div class='vc-body'>
                            {badge_html}
                            <div class='vc-title'>{row['title']}</div>
                            <div class='vc-channel'>📺 {row['channel']} · {uzb_date(row['published'])}</div>
                            <div class='vc-stats'>
                                <div class='vc-stat'><span class='vc-stat-val'>{row['outlier']}x</span><span class='vc-stat-lbl'>Score</span></div>
                                <div class='vc-stat'><span class='vc-stat-val'>{fmt(row['views'])}</span><span class='vc-stat-lbl'>Ko'rishlar</span></div>
                                <div class='vc-stat'><span class='vc-stat-val'>{fmt(row['subs'])}</span><span class='vc-stat-lbl'>Obunachi</span></div>
                            </div>
                        </div>
                    </div>
                    <a href='{row['url']}' target='_blank' style='display:block;text-align:center;
                       margin-top:8px;font-size:12px;color:#6c63ff;text-decoration:none;'>
                       ▶ YouTube da ko'rish</a>
                    """, unsafe_allow_html=True)

            st.markdown("<div class='section-title'>📺 Eng Faol Kanallar</div>", unsafe_allow_html=True)
            ch_stats = df.groupby('channel').agg(
                Videolar=('id','count'),
                Urtacha_Score=('outlier','mean'),
                Jami_Korishlar=('views','sum')
            ).round(1).sort_values('Urtacha_Score',ascending=False).head(10)
            ch_stats['Jami_Korishlar'] = ch_stats['Jami_Korishlar'].apply(fmt)
            fig_ch = px.scatter(ch_stats.reset_index(),x='Videolar',y='Urtacha_Score',
                size='Videolar',color='Urtacha_Score',hover_name='channel',
                color_continuous_scale=['#2a2a58','#6c63ff','#ff4757'],
                labels={'Urtacha_Score':"O'rtacha Viral Score",'Videolar':"Videolar soni"},
                size_max=40)
            fig_ch.update_layout(paper_bgcolor='rgba(0,0,0,0)',plot_bgcolor='rgba(0,0,0,0)',
                height=300,margin=dict(l=0,r=0,t=10,b=0),font=dict(color='#aaaacc',family='Inter'),
                xaxis=dict(gridcolor='#1e1e2e',color='#666688'),
                yaxis=dict(gridcolor='#1e1e2e',color='#666688'),coloraxis_showscale=False)
            st.plotly_chart(fig_ch, use_container_width=True)

# ══════════════════════════════════════════
# TAB 2: VIDEO CARDS
# ══════════════════════════════════════════
with TAB_CARDS:
    need_sub2 = (not st.session_state.authenticated and
                 not is_subscribed(uid) and get_trial(uid)==0)
    if need_sub2:
        show_sub_block("cards")
    elif not st.session_state.results:
        st.info("🔍 Avval chap paneldan qidiruv boshlang.")
    else:
        df = pd.DataFrame(st.session_state.results)
        fc1,fc2,fc3 = st.columns(3)
        sort_by  = fc1.selectbox("Saralash:", ["Viral Score","Ko'rishlar","Layklar"], key="sort_c")
        min_s    = fc2.slider("Min Score:", 0, int(df['outlier'].max()+1), 0, key="min_s")
        cols_cnt = fc3.select_slider("Ustunlar:", [2,3,4], value=3, key="cols_c")
        sort_map = {"Viral Score":"outlier","Ko'rishlar":"views","Layklar":"likes"}
        filtered = df[df['outlier']>=min_s].sort_values(sort_map[sort_by],ascending=False)
        st.caption(f"Ko'rsatilmoqda: {len(filtered)} ta video")
        rows = [filtered.iloc[i:i+cols_cnt] for i in range(0,len(filtered),cols_cnt)]
        for row_df in rows:
            cols = st.columns(cols_cnt)
            for ci,(_, row) in enumerate(row_df.iterrows()):
                with cols[ci]:
                    badge_html = viral_badge(row['outlier'])
                    engage = round((row['likes']+row['comments'])/max(row['views'],1)*100,2)
                    st.markdown(f"""
                    <div class='video-card'>
                        <img class='vc-thumbnail' src='{row['thumbnail']}'
                             onerror="this.src='https://via.placeholder.com/320x180/1a1a2e/6c63ff?text=📺'">
                        <div class='vc-body'>
                            {badge_html}
                            <div class='vc-title'>{row['title']}</div>
                            <div class='vc-channel'>📺 {row['channel']} · {uzb_date(row['published'])}</div>
                            <div class='vc-stats'>
                                <div class='vc-stat'><span class='vc-stat-val'>{row['outlier']}x</span><span class='vc-stat-lbl'>Score</span></div>
                                <div class='vc-stat'><span class='vc-stat-val'>{fmt(row['views'])}</span><span class='vc-stat-lbl'>Ko'rishlar</span></div>
                                <div class='vc-stat'><span class='vc-stat-val'>{engage}%</span><span class='vc-stat-lbl'>Engage</span></div>
                            </div>
                        </div>
                    </div>
                    <a href='{row['url']}' target='_blank'
                       style='display:block;text-align:center;margin:6px 0 14px;
                              font-size:12px;color:#6c63ff;text-decoration:none;
                              padding:6px;background:#1a1a2e;border-radius:8px;'>
                       ▶ YouTube da ko'rish</a>
                    """, unsafe_allow_html=True)

# ══════════════════════════════════════════
# TAB 3: TABLE
# ══════════════════════════════════════════
with TAB_TABLE:
    need_sub3 = (not st.session_state.authenticated and
                 not is_subscribed(uid) and get_trial(uid)==0)
    if need_sub3:
        show_sub_block("table")
    elif not st.session_state.results:
        st.info("🔍 Avval chap paneldan qidiruv boshlang.")
    else:
        df = pd.DataFrame(st.session_state.results)
        st.markdown(f"<div class='section-title'>📊 Jadval <span>{len(df)} ta video</span></div>",
                    unsafe_allow_html=True)
        rows_html = []
        for _, row in df.iterrows():
            engage   = round((row["likes"]+row["comments"])/max(row["views"],1)*100, 2)
            score    = row["outlier"]
            thumb    = html_mod.escape(str(row["thumbnail"]))
            url      = html_mod.escape(str(row["url"]))
            title    = html_mod.escape(str(row["title"])[:65])
            channel  = html_mod.escape(str(row["channel"])[:20])
            sana     = html_mod.escape(uzb_date(str(row["published"])))
            tags_s   = html_mod.escape(str(row.get("tags","—"))[:50])
            est_inc  = row.get("est_income", 0)
            is_mon   = row.get("is_monetized", False)
            mon_txt  = "✅ Ha" if is_mon else "❌ Yo'q"
            mon_clr  = "#2ed573" if is_mon else "#ff4757"
            if score>=100:   sc_color="#ff4757"
            elif score>=50:  sc_color="#ffa502"
            elif score>=20:  sc_color="#6c63ff"
            else:            sc_color="#2ed573"
            rows_html.append(
                f"<tr>"
                f"<td style='padding:8px;'><img src='{thumb}' style='width:80px;height:46px;object-fit:cover;border-radius:6px;'></td>"
                f"<td style='padding:8px 10px;max-width:220px;'><a href='{url}' target='_blank' "
                f"style='color:#9c93ff;text-decoration:none;font-weight:600;font-size:12px;line-height:1.4;'>{title}…</a></td>"
                f"<td style='padding:8px 10px;color:#aaaacc;font-size:12px;'>{channel}</td>"
                f"<td style='padding:8px 10px;color:{sc_color};font-weight:800;font-size:14px;'>{score}x</td>"
                f"<td style='padding:8px 10px;color:#e8e8f0;font-weight:600;'>{fmt(row['views'])}</td>"
                f"<td style='padding:8px 10px;color:#aaaacc;'>{fmt(row['subs'])}</td>"
                f"<td style='padding:8px 10px;color:#aaaacc;'>{fmt(row['likes'])}</td>"
                f"<td style='padding:8px 10px;color:#6c63ff;font-weight:700;'>{engage}%</td>"
                f"<td style='padding:8px 10px;color:{mon_clr};font-weight:700;font-size:13px;'>{mon_txt}</td>"
                f"<td style='padding:8px 10px;color:#2ed573;font-weight:700;'>${est_inc}</td>"
                f"<td style='padding:8px 10px;color:#666688;font-size:11px;max-width:150px;'>{tags_s}</td>"
                f"<td style='padding:8px 10px;color:#555577;font-size:12px;'>{sana}</td>"
                f"</tr>"
            )
        th = "padding:10px 8px;color:#9c93ff;font-size:10px;text-transform:uppercase;letter-spacing:1px;text-align:left;border-bottom:2px solid #2a2a4a;white-space:nowrap;"
        table_html = (
            "<div style='overflow-x:auto;border-radius:14px;border:1px solid #1e1e2e;margin-top:12px;'>"
            "<table style='width:100%;border-collapse:collapse;background:#0f0f1a;font-family:Inter,sans-serif;'>"
            "<thead><tr style='background:#1a1a2e;'>"
            f"<th style='{th}'>🖼</th>"
            f"<th style='{th}'>Sarlavha</th>"
            f"<th style='{th}'>Kanal</th>"
            f"<th style='{th}'>⚡ Score</th>"
            f"<th style='{th}'>👁 Ko\'rish</th>"
            f"<th style='{th}'>👥 Obuna</th>"
            f"<th style='{th}'>👍 Layk</th>"
            f"<th style='{th}'>💬 Engage</th>"
            f"<th style='{th}'>💰 Monetiz</th>"
            f"<th style='{th}'>📈 ~Daromad</th>"
            f"<th style='{th}'>🏷️ Teglar</th>"
            f"<th style='{th}'>📅 Sana</th>"
            "</tr></thead><tbody>" + "".join(rows_html) + "</tbody></table></div>"
        )
        st.markdown(table_html, unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        _topic_name = st.session_state.get("last_topic","result")
        # Excel fayl yaratish va yuklab olish
        _topic_excel = st.session_state.get("last_topic","result")
        try:
            excel_data = create_excel(df, _topic_excel)
            st.download_button(
                "⬇️ Excel yuklab olish (.xlsx)",
                data=excel_data,
                file_name=f"viral777_{_topic_excel}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        except Exception as ex:
            # Fallback: CSV
            csv_fb = df[["title","channel","outlier","views","subs","likes","comments","url"]].to_csv(index=False, encoding="utf-8-sig")
            st.download_button("⬇️ CSV yuklab olish", csv_fb.encode("utf-8-sig"),
                f"viral777_{_topic_excel}_{datetime.now().strftime('%Y%m%d')}.csv",
                "text/csv", use_container_width=True)

# ══════════════════════════════════════════
# TAB 4: CHARTS
# ══════════════════════════════════════════
with TAB_CHART:
    need_sub4 = (not st.session_state.authenticated and
                 not is_subscribed(uid) and get_trial(uid)==0)
    if need_sub4:
        show_sub_block("charts")
    elif not st.session_state.results:
        st.info("🔍 Avval chap paneldan qidiruv boshlang.")
    else:
        df = pd.DataFrame(st.session_state.results)
        ch1, ch2 = st.columns(2)
        with ch1:
            st.markdown("<div class='section-title'>👁 Ko'rishlar vs Viral Score</div>", unsafe_allow_html=True)
            fig1 = px.scatter(df,x='views',y='outlier',size='likes',color='outlier',hover_name='title',
                hover_data={'channel':True,'views':':,.0f','outlier':':.1f'},
                color_continuous_scale=['#2a2a58','#6c63ff','#ff4757'],
                labels={'views':"Ko'rishlar",'outlier':'Viral Score'},size_max=30,log_x=True)
            fig1.update_layout(paper_bgcolor='rgba(0,0,0,0)',plot_bgcolor='rgba(0,0,0,0)',
                height=320,margin=dict(l=0,r=0,t=10,b=0),font=dict(color='#aaaacc',family='Inter'),
                xaxis=dict(gridcolor='#1e1e2e',color='#666688'),
                yaxis=dict(gridcolor='#1e1e2e',color='#666688'),coloraxis_showscale=False)
            st.plotly_chart(fig1, use_container_width=True)
        with ch2:
            st.markdown("<div class='section-title'>📊 Viral Score Distribution</div>", unsafe_allow_html=True)
            fig2 = px.histogram(df,x='outlier',nbins=20,color_discrete_sequence=['#6c63ff'],
                labels={'outlier':'Viral Score','count':"Videolar soni"})
            fig2.update_layout(paper_bgcolor='rgba(0,0,0,0)',plot_bgcolor='rgba(0,0,0,0)',
                height=320,margin=dict(l=0,r=0,t=10,b=0),font=dict(color='#aaaacc',family='Inter'),
                xaxis=dict(gridcolor='#1e1e2e',color='#666688'),
                yaxis=dict(gridcolor='#1e1e2e',color='#666688'),bargap=0.1)
            fig2.update_traces(marker_line_width=0,opacity=0.85)
            st.plotly_chart(fig2, use_container_width=True)
        ch3, ch4 = st.columns(2)
        with ch3:
            st.markdown("<div class='section-title'>📺 Eng Ko'p Video Chiqqan Kanallar</div>", unsafe_allow_html=True)
            ch_cnt = df.groupby('channel').size().sort_values(ascending=False).head(10)
            fig3 = px.bar(x=ch_cnt.values,y=ch_cnt.index,orientation='h',color=ch_cnt.values,
                color_continuous_scale=['#2a2a58','#6c63ff'],labels={'x':"Videolar soni",'y':''})
            fig3.update_layout(paper_bgcolor='rgba(0,0,0,0)',plot_bgcolor='rgba(0,0,0,0)',
                height=320,margin=dict(l=0,r=0,t=10,b=0),font=dict(color='#aaaacc',family='Inter'),
                xaxis=dict(gridcolor='#1e1e2e',color='#666688'),
                yaxis=dict(color='#ccccdd'),coloraxis_showscale=False)
            st.plotly_chart(fig3, use_container_width=True)
        with ch4:
            st.markdown("<div class='section-title'>💬 Engagement Rate Tahlili</div>", unsafe_allow_html=True)
            df['engage'] = ((df['likes']+df['comments'])/df['views'].clip(lower=1)*100).round(2)
            fig4 = px.box(df,y='engage',points='all',color_discrete_sequence=['#6c63ff'],
                labels={'engage':'Engagement Rate (%)'})
            fig4.update_layout(paper_bgcolor='rgba(0,0,0,0)',plot_bgcolor='rgba(0,0,0,0)',
                height=320,margin=dict(l=0,r=0,t=10,b=0),font=dict(color='#aaaacc',family='Inter'),
                yaxis=dict(gridcolor='#1e1e2e',color='#666688'))
            st.plotly_chart(fig4, use_container_width=True)

# ══════════════════════════════════════════
# TAB 5: HISTORY
# ══════════════════════════════════════════
with TAB_HISTORY:
    if st.session_state.history:
        st.markdown("<div class='section-title'>🕐 Qidiruv Tarixi</div>", unsafe_allow_html=True)

        badge_colors = [
            ("#ff4757","rgba(255,71,87,0.15)"),("# ffa502","rgba(255,165,2,0.15)"),
            ("#6c63ff","rgba(108,99,255,0.15)"),("#2ed573","rgba(46,213,115,0.15)"),
            ("#ff6b81","rgba(255,107,129,0.15)"),
        ]
        hist_rows_list = []
        for i, h in enumerate(reversed(st.session_state.history)):
            bg = "#13131f" if i%2==0 else "#0f0f1a"
            bc, bbg = badge_colors[i % len(badge_colors)]
            topildi = h.get("Topildi","-")
            mavzu   = html_mod.escape(str(h.get("Mavzu","-")))
            bozor   = html_mod.escape(str(h.get("Bozor","-")))
            vaqt    = html_mod.escape(str(h.get("Vaqt","-")))
            sana    = html_mod.escape(str(h.get("Sana","-")))
            davr    = html_mod.escape(str(h.get("Davr","-")))
            hist_rows_list.append(
                f"<tr style='background:{bg};border-bottom:1px solid #1a1a2e;'>"
                f"<td style='padding:13px 16px;color:#aaaacc;font-size:13px;white-space:nowrap;'>"
                f"<span style='color:#6c63ff;font-weight:700;'>{sana}</span> "
                f"<span style='color:#555577;'>{vaqt}</span></td>"
                f"<td style='padding:13px 16px;'>"
                f"<span style='background:rgba(108,99,255,0.12);border:1px solid rgba(108,99,255,0.25);"
                f"border-radius:8px;padding:5px 12px;color:#9c93ff;font-weight:800;font-size:14px;'>"
                f"🔍 {mavzu}</span></td>"
                f"<td style='padding:13px 16px;color:#ccccdd;font-size:13px;'>{bozor}</td>"
                f"<td style='padding:13px 16px;color:#888899;font-size:13px;'>{davr}</td>"
                f"<td style='padding:13px 16px;'>"
                f"<span style='background:{bbg};border:1px solid {bc};"
                f"border-radius:10px;padding:5px 14px;color:{bc};font-weight:800;font-size:14px;'>"
                f"{topildi} ta video</span></td>"
                f"</tr>"
            )
        th_st = "padding:13px 16px;color:#9c93ff;font-size:11px;text-transform:uppercase;letter-spacing:1.2px;text-align:left;border-bottom:2px solid #2a2a4a;font-weight:700;"
        hist_html = (
            "<div style='overflow-x:auto;border-radius:16px;border:1px solid #1e1e2e;"
            "margin-top:12px;box-shadow:0 4px 24px rgba(0,0,0,0.4);'>"
            "<table style='width:100%;border-collapse:collapse;background:#0a0a14;font-family:Inter,sans-serif;'>"
            "<thead><tr style='background:#12122a;'>"
            f"<th style='{th_st}'>📅 Sana / Vaqt</th>"
            f"<th style='{th_st}'>🔍 Qidirilgan mavzu</th>"
            f"<th style='{th_st}'>🌍 Bozor</th>"
            f"<th style='{th_st}'>📆 Davr</th>"
            f"<th style='{th_st}'>📊 Natija</th>"
            "</tr></thead><tbody>" + "".join(hist_rows_list) + "</tbody></table></div>"
        )
        st.markdown(hist_html, unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🗑 Tarixni tozalash"):
            st.session_state.history = []
            # Fayldan ham o'chirish
            try:
                hist_file = f"/tmp/v777_hist_{uid[:12]}.json"
                if os.path.exists(hist_file):
                    os.remove(hist_file)
            except: pass
            st.rerun()
    else:
        st.info("🔍 Hali qidiruv amalga oshirilmagan.")

# ══════════════════════════════════════════
# TAB 6: QO'LLANMA
# ══════════════════════════════════════════
with TAB_GUIDE:
    st.markdown("""
    <div style='max-width:820px;margin:0 auto;padding:20px 0;'>

    <div style='text-align:center;margin-bottom:36px;'>
        <div style='font-size:52px;'>📖</div>
        <h2 style='color:#fff;font-size:26px;font-weight:900;margin:10px 0 6px;'>
            Foydalanish Qo\'llanmasi
        </h2>
        <p style='color:#555577;font-size:14px;'>Viral 777 platformasidan to\'liq foydalanish uchun</p>
    </div>

    <!-- API KEY -->
    <div style='background:#0f0f1a;border:1px solid #1e1e2e;border-radius:16px;padding:26px 30px;margin-bottom:20px;'>
        <div style='display:flex;align-items:center;gap:12px;margin-bottom:18px;'>
            <span style='font-size:28px;'>🔑</span>
            <div>
                <h3 style='color:#9c93ff;font-size:18px;font-weight:800;margin:0;'>YouTube API Key olish</h3>
                <p style='color:#555577;font-size:12px;margin:3px 0 0;'>Bepul · Kuniga 10,000 so\'rov · Google hisobi kerak</p>
            </div>
        </div>
        <div style='display:flex;flex-direction:column;gap:10px;'>
            <div style='display:flex;align-items:flex-start;gap:12px;'>
                <span style='background:rgba(108,99,255,0.2);border:1px solid #6c63ff;border-radius:50%;min-width:26px;height:26px;display:flex;align-items:center;justify-content:center;color:#9c93ff;font-weight:800;font-size:12px;'>1</span>
                <div><p style='color:#e8e8f0;font-size:14px;margin:0;font-weight:600;'>Google Cloud Console ga kiring</p>
                <a href="https://console.cloud.google.com" target="_blank" style='color:#6c63ff;font-size:12px;'>→ console.cloud.google.com</a></div>
            </div>
            <div style='display:flex;align-items:flex-start;gap:12px;'>
                <span style='background:rgba(108,99,255,0.2);border:1px solid #6c63ff;border-radius:50%;min-width:26px;height:26px;display:flex;align-items:center;justify-content:center;color:#9c93ff;font-weight:800;font-size:12px;'>2</span>
                <p style='color:#e8e8f0;font-size:14px;margin:0;'>Yuqori qismda <b style="color:#9c93ff;">➕ New Project</b> → nom bering → <b style="color:#9c93ff;">Create</b></p>
            </div>
            <div style='display:flex;align-items:flex-start;gap:12px;'>
                <span style='background:rgba(108,99,255,0.2);border:1px solid #6c63ff;border-radius:50%;min-width:26px;height:26px;display:flex;align-items:center;justify-content:center;color:#9c93ff;font-weight:800;font-size:12px;'>3</span>
                <p style='color:#e8e8f0;font-size:14px;margin:0;'>Chap menyu → <b style="color:#9c93ff;">APIs & Services</b> → <b style="color:#9c93ff;">Library</b> → qidiruv: <b style="color:#ffa502;">YouTube Data API v3</b> → <b style="color:#2ed573;">Enable</b></p>
            </div>
            <div style='display:flex;align-items:flex-start;gap:12px;'>
                <span style='background:rgba(108,99,255,0.2);border:1px solid #6c63ff;border-radius:50%;min-width:26px;height:26px;display:flex;align-items:center;justify-content:center;color:#9c93ff;font-weight:800;font-size:12px;'>4</span>
                <p style='color:#e8e8f0;font-size:14px;margin:0;'><b style="color:#9c93ff;">Credentials</b> → <b style="color:#9c93ff;">+ Create Credentials</b> → <b style="color:#9c93ff;">API Key</b> → Kalitni nusxalang ✅</p>
            </div>
            <div style='display:flex;align-items:flex-start;gap:12px;'>
                <span style='background:rgba(108,99,255,0.2);border:1px solid #6c63ff;border-radius:50%;min-width:26px;height:26px;display:flex;align-items:center;justify-content:center;color:#9c93ff;font-weight:800;font-size:12px;'>5</span>
                <p style='color:#e8e8f0;font-size:14px;margin:0;'>Chap paneldagi <b style="color:#ffa502;">🔑 YouTube API Key</b> bo\'limiga kalitni joylashtiring</p>
            </div>
        </div>
        <div style='background:rgba(255,71,87,0.08);border:1px solid rgba(255,71,87,0.2);border-radius:10px;padding:12px 16px;margin-top:16px;'>
            <p style='color:#ff6b6b;font-size:13px;margin:0;'>⚠️ <b>Muhim:</b> API kalitni hech kim bilan ulashmang! Bepul limit: kuniga <b>10,000 so\'rov</b>. Tugasa ertasi kuni avtomatik tiklanadi.</p>
        </div>
    </div>

    <!-- DASTUR HAQIDA -->
    <div style='background:#0f0f1a;border:1px solid #1e1e2e;border-radius:16px;padding:26px 30px;margin-bottom:20px;'>
        <div style='display:flex;align-items:center;gap:12px;margin-bottom:18px;'>
            <span style='font-size:28px;'>🚀</span>
            <h3 style='color:#9c93ff;font-size:18px;font-weight:800;margin:0;'>Dastur Haqida</h3>
        </div>
        <div style='display:grid;grid-template-columns:1fr 1fr;gap:14px;'>
            <div style='background:#13131f;border-radius:12px;padding:16px;'>
                <div style='font-size:22px;margin-bottom:6px;'>⚡</div>
                <h4 style='color:#fff;font-size:13px;font-weight:700;margin:0 0 5px;'>Viral Score nima?</h4>
                <p style='color:#888899;font-size:12px;margin:0;line-height:1.5;'>Video ko\'rishlari ÷ kanal obunachilari. Masalan: 100,000 ko\'rish ÷ 1,000 obunachi = <b style="color:#ff4757;">100x</b> Mega Viral</p>
            </div>
            <div style='background:#13131f;border-radius:12px;padding:16px;'>
                <div style='font-size:22px;margin-bottom:6px;'>🎯</div>
                <h4 style='color:#fff;font-size:13px;font-weight:700;margin:0 0 5px;'>Qaysi score yaxshi?</h4>
                <p style='color:#888899;font-size:12px;margin:0;line-height:1.6;'>
                    <span style="color:#ff4757;">🔥 100x+</span> = Mega viral<br>
                    <span style="color:#ffa502;">🚀 50–100x</span> = Viral<br>
                    <span style="color:#6c63ff;">⚡ 20–50x</span> = Trendda<br>
                    <span style="color:#2ed573;">✅ 10–20x</span> = Yaxshi
                </p>
            </div>
            <div style='background:#13131f;border-radius:12px;padding:16px;'>
                <div style='font-size:22px;margin-bottom:6px;'>📅</div>
                <h4 style='color:#fff;font-size:13px;font-weight:700;margin:0 0 5px;'>Davr tanlash</h4>
                <p style='color:#888899;font-size:12px;margin:0;line-height:1.5;'>
                    <b style="color:#fff;">7 kun</b> — hozirgi trendlar<br>
                    <b style="color:#fff;">30 kun</b> — oylik tendentsiya<br>
                    <b style="color:#fff;">90+ kun</b> — uzoq muddatli trend
                </p>
            </div>
            <div style='background:#13131f;border-radius:12px;padding:16px;'>
                <div style='font-size:22px;margin-bottom:6px;'>🌍</div>
                <h4 style='color:#fff;font-size:13px;font-weight:700;margin:0 0 5px;'>Bozor tanlash</h4>
                <p style='color:#888899;font-size:12px;margin:0;line-height:1.5;'>
                    <b style="color:#fff;">US/GB</b> — eng katta bozor<br>
                    <b style="color:#fff;">UZ/RU</b> — mahalliy trendlar<br>
                    Bir nishani har bozorda tekshiring!
                </p>
            </div>
        </div>
    </div>

    <!-- TABLAR -->
    <div style='background:#0f0f1a;border:1px solid #1e1e2e;border-radius:16px;padding:26px 30px;margin-bottom:20px;'>
        <div style='display:flex;align-items:center;gap:12px;margin-bottom:16px;'>
            <span style='font-size:28px;'>📑</span>
            <h3 style='color:#9c93ff;font-size:18px;font-weight:800;margin:0;'>Tablar va Funksiyalar</h3>
        </div>
        <div style='display:flex;flex-direction:column;gap:8px;'>
            <div style='display:flex;align-items:center;gap:12px;padding:10px 14px;background:#13131f;border-radius:10px;'>
                <span style='font-size:18px;'>🔥</span>
                <div><b style='color:#fff;font-size:13px;'>Trend Tahlili</b>
                <p style='color:#555577;font-size:11px;margin:2px 0 0;'>Umumiy statistika, viral score ko\'rsatkichi, top 10 grafik, eng viral videolar</p></div>
            </div>
            <div style='display:flex;align-items:center;gap:12px;padding:10px 14px;background:#13131f;border-radius:10px;'>
                <span style='font-size:18px;'>🎬</span>
                <div><b style='color:#fff;font-size:13px;'>Video Kartochkalar</b>
                <p style='color:#555577;font-size:11px;margin:2px 0 0;'>Thumbnail, sarlavha, score, engage% — karta ko\'rinishida</p></div>
            </div>
            <div style='display:flex;align-items:center;gap:12px;padding:10px 14px;background:#13131f;border-radius:10px;'>
                <span style='font-size:18px;'>📊</span>
                <div><b style='color:#fff;font-size:13px;'>Jadval</b>
                <p style='color:#555577;font-size:11px;margin:2px 0 0;'>Barcha ma\'lumotlar jadvali, CSV yuklab olish imkoni (Excel uchun)</p></div>
            </div>
            <div style='display:flex;align-items:center;gap:12px;padding:10px 14px;background:#13131f;border-radius:10px;'>
                <span style='font-size:18px;'>📈</span>
                <div><b style='color:#fff;font-size:13px;'>Grafiklar</b>
                <p style='color:#555577;font-size:11px;margin:2px 0 0;'>Scatter, histogram, kanal tahlili, engagement grafiklari</p></div>
            </div>
            <div style='display:flex;align-items:center;gap:12px;padding:10px 14px;background:#13131f;border-radius:10px;'>
                <span style='font-size:18px;'>🕐</span>
                <div><b style='color:#fff;font-size:13px;'>Tarix</b>
                <p style='color:#555577;font-size:11px;margin:2px 0 0;'>Barcha qidiruvlaringiz tarixi — sana, mavzu, natija</p></div>
            </div>
        </div>
    </div>

    <!-- MASLAHATLAR -->
    <div style='background:linear-gradient(135deg,#0f0f1a,#12121f);border:1px solid #2a2a4a;border-radius:16px;padding:26px 30px;'>
        <div style='display:flex;align-items:center;gap:12px;margin-bottom:16px;'>
            <span style='font-size:28px;'>💡</span>
            <h3 style='color:#9c93ff;font-size:18px;font-weight:800;margin:0;'>Foydali Maslahatlar</h3>
        </div>
        <div style='display:flex;flex-direction:column;gap:8px;'>
            <div style='display:flex;gap:10px;'><span style='color:#2ed573;font-size:14px;flex-shrink:0;'>✓</span>
            <p style='color:#ccccdd;font-size:13px;margin:0;'><b style="color:#fff;">Min Outlier Score</b> ni <b style="color:#ffa502;">50+</b> qo\'ying — faqat haqiqiy viral videolarni ko\'rish uchun</p></div>
            <div style='display:flex;gap:10px;'><span style='color:#2ed573;font-size:14px;flex-shrink:0;'>✓</span>
            <p style='color:#ccccdd;font-size:13px;margin:0;'><b style="color:#fff;">7 kunlik davr + US bozori</b> — eng yangi trendlarni ko\'rish uchun ideal</p></div>
            <div style='display:flex;gap:10px;'><span style='color:#2ed573;font-size:14px;flex-shrink:0;'>✓</span>
            <p style='color:#ccccdd;font-size:13px;margin:0;'>Bir nishani bir necha bozorda tekshiring — <b style="color:#fff;">US, GB, UZ</b> farqli trendlar ko\'rsatadi</p></div>
            <div style='display:flex;gap:10px;'><span style='color:#2ed573;font-size:14px;flex-shrink:0;'>✓</span>
            <p style='color:#ccccdd;font-size:13px;margin:0;'><b style="color:#fff;">Engage% 2%+</b> bo\'lgan videolar — auditoriya juda faol nisha</p></div>
            <div style='display:flex;gap:10px;'><span style='color:#6c63ff;font-size:14px;flex-shrink:0;'>💡</span>
            <p style='color:#ccccdd;font-size:13px;margin:0;'>Natijalarni <b style="color:#9c93ff;">CSV</b> sifatida yuklab, Excel da chuqurroq tahlil qiling</p></div>
        </div>
    </div>

    </div>
    """, unsafe_allow_html=True)


# ══════════════════════════════════════════
# TAB 7: ADMIN PANEL
# ══════════════════════════════════════════
with TAB_ADMIN:
    if not st.session_state.authenticated:
        st.warning("🔒 Bu bo\'lim faqat Admin uchun!")
        st.stop()

    st.markdown("<div class=\'section-title\'>🛡️ Admin Boshqaruvi</div>", unsafe_allow_html=True)

    db  = load_db()
    now = datetime.now()

    # Foydalanuvchilarni ajratish
    all_users    = {k:v for k,v in db.items()
                    if k not in ("activation_codes","login_codes")
                    and isinstance(v, dict) and "subscribed" in v}
    codes        = db.get("activation_codes", {})
    login_codes  = db.get("login_codes", {})

    active_users = {k:v for k,v in all_users.items()
                    if v.get("subscribed") and v.get("sub_until") and
                    datetime.fromisoformat(v["sub_until"]) > now}
    trial_users  = {k:v for k,v in all_users.items()
                    if not v.get("subscribed") and v.get("trial_used",0) > 0}
    used_codes   = {k:v for k,v in codes.items() if v.get("used")}
    unused_codes = {k:v for k,v in codes.items() if not v.get("used")}

    # ── KPI kartalar ──
    k1,k2,k3,k4,k5 = st.columns(5)
    for col, label, val, color in [
        (k1, "Jami Mijozlar",   len(all_users),    "purple"),
        (k2, "Faol Obuna",      len(active_users),  "green"),
        (k3, "Sinov Ishlatgan", len(trial_users),   "gold"),
        (k4, "Ishlatilgan Kod", len(used_codes),    "red"),
        (k5, "Kutayotgan Kod",  len(unused_codes),  "purple"),
    ]:
        col.markdown(
            f"<div class=\'stat-card {color}\'>"
            f"<div class=\'stat-label\'>{label}</div>"
            f"<div class=\'stat-value\'>{val}</div></div>",
            unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Faol obunalar ──
    st.markdown("<div class=\'section-title\'>✅ Faol Obunalar</div>", unsafe_allow_html=True)
    if active_users:
        rows = []
        for uk, ud in active_users.items():
            sub_dt    = datetime.fromisoformat(ud["sub_until"])
            days_left = (sub_dt - now).days
            tg_id     = ud.get("telegram_id", uk[:14]+"...")
            orders    = ud.get("orders", [])
            last_ord  = orders[-1].get("date","—")[:10] if orders else "—"
            plan      = "1 yil" if days_left>300 else ("3 oy" if days_left>60 else "1 oy")
            rows.append({
                "TG / UID":     tg_id,
                "Tugaydi":      sub_dt.strftime("%d.%m.%Y"),
                "Qoldi (kun)":  days_left,
                "Reja":         plan,
                "Sotib olgan":  last_ord,
                "Sinov":        str(ud.get("trial_used",0))+"/3",
            })
        st.dataframe(
            pd.DataFrame(rows).sort_values("Qoldi (kun)"),
            use_container_width=True, hide_index=True)
    else:
        st.info("Hozircha faol obuna yo\'q.")

    st.divider()

    # ── Barcha foydalanuvchilar ──
    st.markdown("<div class=\'section-title\'>👥 Barcha Foydalanuvchilar</div>", unsafe_allow_html=True)
    all_rows = []
    for uk, ud in all_users.items():
        sub_str   = "—"
        days_left = 0
        status    = "❌ Yo\'q"
        if ud.get("subscribed") and ud.get("sub_until"):
            sub_dt    = datetime.fromisoformat(ud["sub_until"])
            days_left = (sub_dt - now).days
            sub_str   = sub_dt.strftime("%d.%m.%Y")
            status    = "✅ Faol" if days_left > 0 else "⏰ Tugagan"
        elif ud.get("trial_used",0) > 0:
            status = "🎁 Sinov (" + str(ud.get("trial_used",0)) + "/3)"
        tg_id = ud.get("telegram_id", uk[:14]+"...")
        all_rows.append({
            "TG / UID":   tg_id,
            "Holat":      status,
            "Tugaydi":    sub_str,
            "Kun qoldi":  max(0, days_left),
            "To\'lovlar": len(ud.get("orders",[])),
            "Sinov":      str(ud.get("trial_used",0))+"/3",
        })
    if all_rows:
        st.dataframe(pd.DataFrame(all_rows), use_container_width=True, hide_index=True)
    else:
        st.info("Hali hech kim ro\'yxatdan o\'tmagan.")

    st.divider()

    # ── Aktivatsiya kodlari ──
    st.markdown("<div class=\'section-title\'>🔑 Aktivatsiya Kodlari</div>", unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    c1.metric("Ishlatilgan", len(used_codes))
    c2.metric("Kutayotgan",  len(unused_codes))

    if unused_codes:
        code_rows = []
        for code, cdata in unused_codes.items():
            exp = datetime.fromisoformat(cdata["expires"])
            ql  = (exp - now).days
            code_rows.append({
                "Kod":     code,
                "TG ID":   cdata.get("telegram_id","—"),
                "Tugaydi": exp.strftime("%d.%m.%Y %H:%M"),
                "Qoldi":   str(ql)+" kun" if ql > 0 else "Tugagan",
            })
        st.dataframe(pd.DataFrame(code_rows), use_container_width=True, hide_index=True)

    st.divider()

    # ── Telegram login tarixi ──
    if login_codes:
        st.markdown("<div class=\'section-title\'>✈️ Telegram Login Tarixi</div>", unsafe_allow_html=True)
        lg_rows = []
        for code, cdata in list(login_codes.items())[-20:]:
            lg_rows.append({
                "TG ID":    cdata.get("tg_id","—"),
                "Ism":      cdata.get("tg_name","—"),
                "Vaqt":     cdata.get("created","—")[:16].replace("T"," "),
                "Holat":    "✅ Ishlatilgan" if cdata.get("used") else "⏳ Kutilmoqda",
            })
        if lg_rows:
            st.dataframe(pd.DataFrame(lg_rows[::-1]), use_container_width=True, hide_index=True)
        st.divider()

    # ── Qo'lda kod yaratish ──
    st.markdown("<div class=\'section-title\'>⚡ Qo\'lda Kod Yaratish</div>", unsafe_allow_html=True)
    m1, m2, m3 = st.columns(3)
    m_note = m1.text_input("Izoh (isim/chek):", key="adm_note")
    m_days = m2.number_input("Muddat (kun):", min_value=1, max_value=365, value=30, key="adm_days")
    m_tg   = m3.text_input("TG ID (ixtiyoriy):", key="adm_tg")

    if st.button("🔑 Yangi Kod Yaratish", use_container_width=True, key="adm_gen"):
        new_code = ""
        existing = db.get("activation_codes", {})
        for _ in range(100):
            c = "".join(random.choices(string.ascii_uppercase + string.digits, k=6))
            if c not in existing:
                new_code = c
                break
        if new_code:
            db.setdefault("activation_codes", {})[new_code] = {
                "telegram_id": m_tg or "manual",
                "order_id":    f"MANUAL_{uuid.uuid4().hex[:6].upper()}",
                "note":        m_note,
                "created":     now.isoformat(),
                "expires":     (now + timedelta(days=int(m_days))).isoformat(),
                "used":        False,
            }
            save_db(db)
            st.success(f"✅ Yangi kod: **`{new_code}`** — {m_days} kun")
            st.rerun()
        else:
            st.error("❌ Kod yaratib bo\'lmadi")

    # ── DB ni yangilash ──
    if st.button("🔄 Ma\'lumotlarni Yangilash", use_container_width=True, key="adm_refresh"):
        st.rerun()
